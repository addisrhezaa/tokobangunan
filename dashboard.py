import tkinter as tk
from tkinter import *
from tkinter import messagebox
from tkinter import ttk
from tkcalendar import DateEntry
import customtkinter
import os
from openpyxl import Workbook, load_workbook
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg 
from PIL import Image
import sqlite3
from datetime import datetime, timedelta

# Database connection
conn = sqlite3.connect('bangunandb.sqlite')
c = conn.cursor()

def center_window(window, width, height):
    # Mendapatkan lebar dan tinggi layar
    screen_width = window.winfo_screenwidth()
    screen_height = window.winfo_screenheight()

    # Menghitung posisi x dan y untuk jendela agar berada di tengah layar
    x = int((screen_width/2) - (width/2))
    y = int((screen_height/2) - (height/2))

    # Menentukan posisi jendela
    root.geometry(f"{width}x{height}+{x}+{y}")
    root.resizable(False,False)

def center_content(frame):
    # Mengatur properti grid agar konten berada di tengah
    frame.grid_configure(padx=10, pady=10)

    # Mengatur properti grid untuk seluruh widget di dalam frame
    for widget in frame.winfo_children():
        widget.grid_configure(padx=5, pady=5)

def format_currency(value):
    formatted_value = 'Rp {:,.2f}'.format(abs(float(value)))
    if value < 0:
        formatted_value = '(' + formatted_value[0:] + ')'
    return formatted_value

def show_dashboard():
    # Hapus konten sebelumnya (jika ada)
    for widget in content_frame.winfo_children():
        widget.destroy()

   # Function to create the chart for product sales
    def create_sales_chart():
        c.execute("SELECT nama_produk, SUM(jumlah) as jumlah_penjualan FROM Produk JOIN DetailJual USING(id_produk) GROUP BY nama_produk")
        data = c.fetchall()
        products = [row[0] for row in data]
        sales = [row[1] for row in data]

        fig = Figure(figsize=(14, 6), dpi=65)
        ax = fig.add_subplot(111)
        
        ax.bar(products, sales)
        ax.set_title("Kuantiti Penjualan").set_fontsize(20)
        ax.set_xlabel("Produk")
        ax.set_ylabel("Jumlah Penjualan")
        ax.set_xticks(range(len(products)))
        ax.set_xticklabels(products, rotation=20, ha='right')

        canvas = FigureCanvasTkAgg(fig, master=content_frame)
        canvas.get_tk_widget().grid(row=0, column=0, columnspan=2, padx=1, pady=1)

    def create_top10_sales_chart():
        c.execute("SELECT nama_produk, SUM(jumlah) as jumlah_penjualan FROM Produk JOIN DetailJual USING(id_produk) GROUP BY nama_produk ORDER BY jumlah_penjualan DESC LIMIT 10")
        data = c.fetchall()
        products = [row[0] for row in data]
        sales = [row[1] for row in data]

        fig = Figure(figsize=(9, 9), dpi=53)
        ax = fig.add_subplot(111)

        ax.bar(products, sales)
        ax.set_title("Top 10 Produk Terlaris").set_fontsize(20)
        ax.set_xlabel("Produk")
        ax.set_ylabel("Jumlah Penjualan")
        ax.set_xticks(range(len(products)))
        ax.set_xticklabels(products, rotation=20, ha='right')

        canvas = FigureCanvasTkAgg(fig, master=content_frame)
        canvas.get_tk_widget().grid(row=1, column=0)

    def create_top10_profit_chart():
        c.execute('''SELECT Produk.nama_produk, SUM(DetailJual.harga_jual * DetailJual.jumlah) AS total_profit
             FROM Produk
             INNER JOIN DetailJual ON Produk.id_produk = DetailJual.id_produk
             GROUP BY Produk.nama_produk
             ORDER BY total_profit DESC
             LIMIT 10''')
        data = c.fetchall()
        products = [row[0] for row in data]
        sales = [row[1] for row in data]

        fig = Figure(figsize=(9, 9), dpi=53)
        ax = fig.add_subplot(111)

        ax.bar(products, sales)
        ax.set_title("Top 10 Produk Menguntungkan").set_fontsize(20)
        ax.set_xlabel("Produk")
        ax.set_ylabel("Keuntungan")
        ax.set_xticks(range(len(products)))
        ax.set_xticklabels(products, rotation=20, ha='right')

        canvas = FigureCanvasTkAgg(fig, master=content_frame)
        canvas.get_tk_widget().grid(row=1, column=1)

    def create_plot_line():
        end_date = datetime.now()
        start_date = end_date - timedelta(days=30)
        
        c.execute("SELECT tanggal_jual, pembayaran FROM Penjualan WHERE tanggal_jual BETWEEN ? AND ?", (start_date, end_date))
        rows = c.fetchall()

        # Prepare the data for the chart
        dates = []
        sales = []
        for row in rows:
            tanggal_jual = datetime.strptime(row[0], "%Y-%m-%d")
            pembayaran = row[1]
            dates.append(tanggal_jual)
            sales.append(pembayaran)

        # Create the chart figure
        fig = Figure(figsize=(14, 6),dpi=65)
        ax = fig.add_subplot(111)
    
        ax.plot(dates, sales, marker='o')
        ax.set_xlabel('Tanggal')
        ax.set_ylabel('Pembayaran')
        ax.set_title('Grafik Tren Penjualan 1 Bulan Terakhir').set_fontsize(20)

        # Create the chart canvas
        canvas = FigureCanvasTkAgg(fig, master=content_frame)
        canvas.get_tk_widget().grid(row=3, column=0, columnspan=2, padx=5, pady=5)

    create_sales_chart()
    create_top10_sales_chart()
    create_top10_profit_chart()
    create_plot_line()

def show_pembelian():
    # Hapus konten sebelumnya (jika ada)
    for widget in content_frame.winfo_children():
        widget.destroy()

    # Function to refresh the table view
    def refresh_table():
        # Clear existing table data
        for row in treeview.get_children():
            treeview.delete(row)
        
        # Fetch and display data from the database
        c.execute("SELECT * FROM Pembelian")
        rows = c.fetchall()
        for row in rows:
            pembelian_id = row[0]
            c.execute("SELECT id_supplier, nama_supplier FROM Supplier")
            supplier_rows = c.fetchall()
            for rowss in supplier_rows:
                if row[1] == rowss[0]:
                    nama_supplier = rowss[1]
            tanggal_transaksi = row[2]
            tanggal_tempo = row[3]
            pembayaran = row[4]
            
            # Insert data into the table view
            treeview.insert("", tk.END, values=(pembelian_id, nama_supplier, tanggal_transaksi, tanggal_tempo, pembayaran))

    # Function to add a new penjualan record
    def add_pembelian():
        # Fetch values from the entry fields

        # Get the selected ID from the combobox
        selected_supplier_id = entry_nama.get()
        c.execute("SELECT id_supplier, nama_supplier FROM Supplier")
        supplier_rows = c.fetchall()
        for row in supplier_rows:
            if selected_supplier_id == row[1]:
                supplier_id = row[0]

        tanggal_transaksi = entry_tanggal.get()
        tanggal_tempo = entry_tanggal_tempo.get()
        pembayaran = entry_pembayaran.get()
        
        # Insert new penjualan record into the database
        c.execute("INSERT INTO Pembelian (id_supplier, tanggal_beli, tanggal_tempo, pembayaran) VALUES (?, ?, ?, ?)",
                (supplier_id, tanggal_transaksi, tanggal_tempo, pembayaran))
        conn.commit()
        
        # Refresh the table view
        refresh_table()

    # Function to delete a penjualan record
    def delete_pembelian():
        # Get the selected item from the table view
        selected_item = treeview.selection()
        if selected_item:
            # Get the penjualan_id from the selected item
            pembelian_id = treeview.item(selected_item)["values"][0]
            
            # Delete the penjualan record from the database
            c.execute("DELETE FROM Pembelian WHERE id_pembelian=?", (pembelian_id,))
            c.execute("DELETE FROM DetailBeli WHERE id_pembelian=?", (pembelian_id,))
            conn.commit()
            
            # Refresh the table view
            refresh_table()
            refresh_detailbeli_table(pembelian_id)

    # Function to update a penjualan record
    def update_pembelian():
        # Get the selected item from the table view
        selected_item = treeview.focus()
        if selected_item:
            # Get the penjualan_id from the selected item
            pembelian_id = treeview.item(selected_item)["values"][0]

            selected_supplier_id = entry_nama.get()
            c.execute("SELECT id_supplier, nama_supplier FROM Supplier")
            supplier_rows = c.fetchall()
            for row in supplier_rows:
                if selected_supplier_id == row[1]:
                    updated_supplier = row[0]

            updated_tanggal = entry_tanggal.get()
            updated_tanggal_tempo = entry_tanggal_tempo.get()
            updated_pembayaran = entry_pembayaran.get()

            c.execute("UPDATE Pembelian SET id_supplier=?, tanggal_beli=?, tanggal_tempo=?, pembayaran=? WHERE id_pembelian=?", 
                    (updated_supplier, updated_tanggal, updated_tanggal_tempo, updated_pembayaran, pembelian_id))

            # Commit the changes to the database
            conn.commit()

            # Refresh the table view
            refresh_table()

    # Function to clear the entry fields
    def clear_fields():
        entry_nama.set("")
        entry_tanggal.delete(0, tk.END)
        entry_pembayaran.delete(0, tk.END)

    def populate_supplier_dropdown():
        c.execute("SELECT id_supplier, nama_supplier FROM Supplier")
        supplier_rows = c.fetchall()

        listsupplier = []
        listidsupp = []
        for row in supplier_rows:
            idsupp = row[0]
            namasupp = row [1]
            listidsupp.append(idsupp)
            listsupplier.append(namasupp)
        
        return  listsupplier

    listsupplier = populate_supplier_dropdown()

    # Create the form widgets
    label_nama = customtkinter.CTkLabel(content_frame, text="Nama Supplier", text_color="black")
    entry_nama = customtkinter.CTkComboBox(content_frame, values=listsupplier)

    label_tanggal = customtkinter.CTkLabel(content_frame, text="Tanggal Transaksi", text_color="black")
    entry_tanggal = DateEntry(content_frame, width=19 , background='darkblue', foreground='white', date_pattern='yyyy-mm-dd')

    label_tanggal_tempo = customtkinter.CTkLabel(content_frame, text="Tanggal Jatuh Tempo", text_color="black")
    entry_tanggal_tempo = DateEntry(content_frame, width=19 , background='darkblue', foreground='white', date_pattern='yyyy-mm-dd')

    label_pembayaran = customtkinter.CTkLabel(content_frame, text="Pembayaran", text_color="black")
    entry_pembayaran = customtkinter.CTkEntry(content_frame)

    button_add = customtkinter.CTkButton(content_frame, text="Add", command=add_pembelian, width=100)
    button_delete = customtkinter.CTkButton(content_frame, text="Delete", command=delete_pembelian, width=100)
    button_update = customtkinter.CTkButton(content_frame, text="Update ", command=update_pembelian, width=100)
    button_clear = customtkinter.CTkButton(content_frame, text="Clear", command=clear_fields, width=100)

    # Position the form widgets
    label_nama.grid(row=0, column=0, padx=5, pady=5)
    entry_nama.grid(row=0, column=1, padx=5, pady=5)

    label_tanggal.grid(row=1, column=0, padx=5, pady=5)
    entry_tanggal.grid(row=1, column=1, padx=5, pady=5)

    label_tanggal_tempo.grid(row=1, column=2, padx=5, pady=5)
    entry_tanggal_tempo.grid(row=1, column=3, padx=5, pady=5)

    label_pembayaran.grid(row=2, column=0, padx=5, pady=5)
    entry_pembayaran.grid(row=2, column=1, padx=5, pady=5)

    button_add.grid(row=3, column=0, padx=5, pady=5)
    button_delete.grid(row=3, column=1, padx=5, pady=5)
    button_update.grid(row=3, column=2, padx=5, pady=5)
    button_clear.grid(row=3, column=3, padx=5, pady=5)

    # Create the table view
    treeview = ttk.Treeview(content_frame, columns=("ID", "Nama Supplier", "Tanggal Transaksi", "Tanggal Jatuh Tempo", "Pembayaran"), show="headings")
    treeview.heading("ID", text="No. Pembelian")
    treeview.heading("Nama Supplier", text="Nama Supplier")
    treeview.heading("Tanggal Transaksi", text="Tanggal Transaksi")
    treeview.heading("Tanggal Jatuh Tempo", text="Tanggal Jatuh Tempo")
    treeview.heading("Pembayaran", text="Pembayaran")

    treeview.column("ID", width=180)
    treeview.column("Nama Supplier", width=180)
    treeview.column("Tanggal Transaksi", width=180)
    treeview.column("Tanggal Jatuh Tempo", width=180)
    treeview.column("Pembayaran", width=180)

    treeview.grid(row=4, column=0, columnspan=4, padx=1, pady=1)

    def on_select(event):
        selected_item = treeview.focus()
        if selected_item:
            pembelian_id, supplier_id, tanggal_transaksi, tanggal_tempo, pembayaran = treeview.item(selected_item, "values")
            entry_nama.set(supplier_id)
            entry_tanggal.delete(0, tk.END)
            entry_tanggal.insert(tk.END, tanggal_transaksi)
            entry_tanggal_tempo.delete(0, tk.END)
            entry_tanggal_tempo.insert(tk.END, tanggal_tempo)
            entry_pembayaran.delete(0, tk.END)
            entry_pembayaran.insert(tk.END, pembayaran)
            
            # Refresh the DetailJual table view based on the selected Penjualan
            refresh_detailbeli_table(pembelian_id)

    treeview.bind("<<TreeviewSelect>>", on_select)

    # Function to add a new DetailJual record
    def add_detailbeli():
        # Get the selected Penjualan ID
        selected_item = treeview.focus()
        if not selected_item:
            messagebox.showerror("Error", "No Pembelian record selected")
            return
        pembelian_id = treeview.item(selected_item, "values")[0]

        # Get the selected ID from the combobox
        selected_produk_id = produk_dropdown.get()
        c.execute("SELECT id_produk, nama_produk FROM Produk")
        produk_rows = c.fetchall()
        for row in produk_rows:
            if selected_produk_id == row[1]:
                produk_id = row[0]
        
        # Get the input from the entry field
        harga_beli = entry_harga_beli.get()
        jumlah_beli = entry_jumlah.get()
        
        # Insert the DetailJual data into the database
        c.execute("INSERT INTO DetailBeli (id_pembelian, id_produk, harga_beli, jumlah_produk) VALUES (?, ?, ?, ?)",
                (pembelian_id, produk_id, harga_beli, jumlah_beli))
        conn.commit()
        
        # Refresh the DetailJual table view
        refresh_detailbeli_table(pembelian_id)

    # Function to update a DetailJual record
    def update_detailbeli():
        selected_item = detailbeli_treeview.focus()
        if selected_item:
            # Get the detailjual_id from the selected item
            detailbeli_id = detailbeli_treeview.item(selected_item)["values"][0]

            # Get the selected ID from the combobox
            selected_produk_id = produk_dropdown.get()
            c.execute("SELECT id_produk, nama_produk FROM Produk")
            produk_rows = c.fetchall()
            for row in produk_rows:
                if selected_produk_id == row[1]:
                    updated_produk = row[0]
            
            updated_harga = entry_harga_beli.get()
            updated_jumlah = entry_jumlah.get()

            # Update the DetailJual record in the database
            c.execute("UPDATE DetailBeli SET id_produk=?, harga_beli=?, jumlah_bahan=? WHERE id_detailbeli=?", 
                    (updated_produk, updated_harga, updated_jumlah, detailbeli_id))
            conn.commit()
            
            # Refresh the DetailJual table view
            pembelian_id = treeview.item(treeview.selection(), "values")[0]
            refresh_detailbeli_table(pembelian_id)
            
    # Function to delete a DetailJual record
    def delete_detailbeli():
        selected_item = detailbeli_treeview.selection()
        if selected_item:
            # Get the detailjual_id from the selected item
            detailbeli_id = detailbeli_treeview.item(selected_item)["values"][0]
            
            # Delete the DetailJual record from the database
            c.execute("DELETE FROM DetailBeli WHERE id_detailbeli=?", (detailbeli_id,))
            conn.commit()
            
            # Refresh the DetailJual table view
            pembelian_id = treeview.item(treeview.selection(), "values")[0]
            refresh_detailbeli_table(pembelian_id)

    def refresh_detailbeli_table(pembelian_id):
        # Clear existing table data
        for row in detailbeli_treeview.get_children():
            detailbeli_treeview.delete(row)
        
        # Fetch and display data from the database
        c.execute("SELECT * FROM DetailBeli WHERE id_pembelian=?", (pembelian_id,))
        rows = c.fetchall()
        for row in rows:
            id_detailjual = row[0]
            c.execute("SELECT id_produk, nama_produk FROM Produk")
            produk_rows = c.fetchall()
            for rowss in produk_rows:
                if row[2] == rowss[0]:
                    nama_produk = rowss[1]
            harga_beli = row[3]
            jumlah = row[4]
            
            # Insert data into the table view
            detailbeli_treeview.insert("", tk.END, values=(id_detailjual, nama_produk, harga_beli, jumlah))

    # Create the DetailJual table view
    detailbeli_treeview = ttk.Treeview(content_frame, columns=("ID Detail Beli", "Nama Produk", "Harga Beli", "Jumlah"), show="headings")
    detailbeli_treeview.heading("ID Detail Beli", text="No. Detail Beli")
    detailbeli_treeview.heading("Nama Produk", text="Nama Produk")
    detailbeli_treeview.heading("Harga Beli", text="Harga Beli")
    detailbeli_treeview.heading("Jumlah", text="Jumlah")
    detailbeli_treeview.column("ID Detail Beli", width=180)
    detailbeli_treeview.column("Nama Produk", width=180)
    detailbeli_treeview.column("Harga Beli", width=180)
    detailbeli_treeview.column("Jumlah", width=180)

    detailbeli_treeview.grid(row=9, column=0, columnspan=4, padx=5, pady=5)

    def on_select_detailbeli(event):
        selected_item = detailbeli_treeview.focus()
        if selected_item:
            detailbeli_id, nama_produk, harga_beli, jumlah = detailbeli_treeview.item(selected_item, "values")
            entry_jumlah.delete(0, tk.END)
            entry_jumlah.insert(tk.END, jumlah)
            entry_harga_beli.delete(0, tk.END)
            entry_harga_beli.insert(tk.END, harga_beli)
            produk_dropdown.set(nama_produk)

    detailbeli_treeview.bind("<<TreeviewSelect>>", on_select_detailbeli)

    def clear_detailbeli_fields():
        entry_jumlah.delete(0, tk.END)
        entry_harga_beli.delete(0, tk.END)
        produk_dropdown.set("")  # Clear the selection

    # Create the entry fields for DetailJual
    # Function to fetch and populate the Produk dropdown
    def populate_produk_dropdown():
        c.execute("SELECT id_produk, nama_produk FROM Produk")
        produk_rows = c.fetchall()

        listproduk = []
        listidprod = []
        for row in produk_rows:
            idprod = row[0]
            namaprod = row [1]
            listidprod.append(idprod)
            listproduk.append(namaprod)
        
        # produk_dropdown['values'] = listproduk
        return listproduk

    listproduk = populate_produk_dropdown()

    # Create the Produk dropdown
    produk_label = customtkinter.CTkLabel(content_frame, text="Nama Produk", text_color="black")
    produk_dropdown = customtkinter.CTkComboBox(content_frame, values=listproduk)
    produk_label.grid(row=5, column=0, padx=5, pady=5)
    produk_dropdown.grid(row=5, column=1, padx=5, pady=5)

    label_harga_beli = customtkinter.CTkLabel(content_frame, text="Harga Beli", text_color="black")
    entry_harga_beli = customtkinter.CTkEntry(content_frame)

    label_harga_beli.grid(row=6, column=0, padx=5, pady=5)
    entry_harga_beli.grid(row=6, column=1, padx=5, pady=5)

    label_jumlah = customtkinter.CTkLabel(content_frame, text="Jumlah", text_color="black")
    entry_jumlah = customtkinter.CTkEntry(content_frame)

    label_jumlah.grid(row=7, column=0, padx=5, pady=5)
    entry_jumlah.grid(row=7, column=1, padx=5, pady=5)

    # Create the button for adding DetailJual
    button_add_detailjual = customtkinter.CTkButton(content_frame, text="Add", command=add_detailbeli, width=100)
    button_add_detailjual.grid(row=8, column=0, padx=5, pady=5)

    # Create the buttons for updating and deleting DetailJual
    button_update_detailjual = customtkinter.CTkButton(content_frame, text="Update", command=update_detailbeli, width=100)
    button_delete_detailjual = customtkinter.CTkButton(content_frame, text="Delete", command=delete_detailbeli, width=100)
    button_clear_detailjual = customtkinter.CTkButton(content_frame, text="Clear", command=clear_detailbeli_fields, width=100)
    button_update_detailjual.grid(row=8, column=2, padx=5, pady=5)
    button_delete_detailjual.grid(row=8, column=1, padx=5, pady=5)
    button_clear_detailjual.grid(row=8, column=3, padx=5, pady=5)

    # Fetch and display initial data in the table view
    refresh_table()

def show_penjualan():
    # Hapus konten sebelumnya (jika ada)
    for widget in content_frame.winfo_children():
        widget.destroy()

    # Function to refresh the table view
    def refresh_table():
        # Clear existing table data
        for row in treeview.get_children():
            treeview.delete(row)
        
        # Fetch and display data from the database
        c.execute("SELECT * FROM Penjualan")
        rows = c.fetchall()
        for row in rows:
            penjualan_id = row[0]
            nama_pembeli = row[1]
            tanggal_transaksi = row[2]
            tanggal_tempo = row[3]
            pembayaran = row[4]
            
            # Insert data into the table view
            treeview.insert("", tk.END, values=(penjualan_id, nama_pembeli, tanggal_transaksi, tanggal_tempo, pembayaran))

    # Function to add a new penjualan record
    def add_penjualan():
        # Fetch values from the entry fields
        nama_pembeli = entry_nama.get()
        tanggal_transaksi = entry_tanggal.get()
        tanggal_tempo = entry_tanggal_tempo.get()
        pembayaran = entry_pembayaran.get()
        
        # Insert new penjualan record into the database
        c.execute("INSERT INTO Penjualan (nama_pembeli, tanggal_jual, tanggal_tempo, pembayaran) VALUES (?, ?, ?, ?)",
                (nama_pembeli, tanggal_transaksi, tanggal_tempo, pembayaran))
        conn.commit()
        
        # Refresh the table view
        refresh_table()

    # Function to delete a penjualan record
    def delete_penjualan():
        # Get the selected item from the table view
        selected_item = treeview.selection()
        if selected_item:
            # Get the penjualan_id from the selected item
            penjualan_id = treeview.item(selected_item)["values"][0]
            
            # Delete the penjualan record from the database
            c.execute("DELETE FROM Penjualan WHERE id_penjualan=?", (penjualan_id,))
            c.execute("DELETE FROM DetailJual WHERE id_penjualan=?", (penjualan_id,))
            conn.commit()
            
            # Refresh the table view
            refresh_table()
            refresh_detailjual_table(penjualan_id)

    # Function to update a penjualan record
    def update_penjualan():
        # Get the selected item from the table view
        selected_item = treeview.focus()
        if selected_item:
            # Get the penjualan_id from the selected item
            penjualan_id = treeview.item(selected_item)["values"][0]

            # Perform database update based on modified values
            updated_nama = entry_nama.get()
            updated_tanggal = entry_tanggal.get()
            updated_tanggal_tempo = entry_tanggal_tempo.get()
            updated_pembayaran = entry_pembayaran.get()

            c.execute("UPDATE Penjualan SET nama_pembeli=?, tanggal_jual=?, tanggal_tempo=?, pembayaran=? WHERE id_penjualan=?", 
                    (updated_nama, updated_tanggal, updated_tanggal_tempo, updated_pembayaran, penjualan_id))

            # Commit the changes to the database
            conn.commit()

            # Refresh the table view
            refresh_table()

    # Function to clear the entry fields
    def clear_fields():
        entry_nama.delete(0, tk.END)
        entry_tanggal.delete(0, tk.END)
        entry_tanggal_tempo.delete(0, tk.END)
        entry_pembayaran.delete(0, tk.END)

    # Create the form widgets
    label_nama = customtkinter.CTkLabel(content_frame, text="Nama Pembeli", text_color="black")
    entry_nama = customtkinter.CTkEntry(content_frame)

    label_tanggal = customtkinter.CTkLabel(content_frame, text="Tanggal Transaksi", text_color="black")
    entry_tanggal = DateEntry(content_frame, width=19, background='darkblue', foreground='white', date_pattern='yyyy-mm-dd')

    label_tanggal_tempo = customtkinter.CTkLabel(content_frame, text="Tanggal Jatuh Tempo", text_color="black")
    entry_tanggal_tempo = DateEntry(content_frame, width=19, background='darkblue', foreground='white', date_pattern='yyyy-mm-dd')

    label_pembayaran = customtkinter.CTkLabel(content_frame, text="Pembayaran", text_color="black")
    entry_pembayaran = customtkinter.CTkEntry(content_frame)

    button_add = customtkinter.CTkButton(content_frame, text="Add", command=add_penjualan, width=100)
    button_delete = customtkinter.CTkButton(content_frame, text="Delete", command=delete_penjualan, width=100)
    button_update = customtkinter.CTkButton(content_frame, text="Update ", command=update_penjualan, width=100)
    button_clear = customtkinter.CTkButton(content_frame, text="Clear", command=clear_fields, width=100)

    # Position the form widgets
    label_nama.grid(row=0, column=0, padx=5, pady=5)
    entry_nama.grid(row=0, column=1, padx=5, pady=5)

    label_tanggal.grid(row=1, column=0, padx=5, pady=5)
    entry_tanggal.grid(row=1, column=1, padx=5, pady=5)

    label_tanggal_tempo.grid(row=1, column=2, padx=5, pady=5)
    entry_tanggal_tempo.grid(row=1, column=3, padx=5, pady=5)

    label_pembayaran.grid(row=2, column=0, padx=5, pady=5)
    entry_pembayaran.grid(row=2, column=1, padx=5, pady=5)

    button_add.grid(row=3, column=0, padx=5, pady=5)
    button_delete.grid(row=3, column=1, padx=5, pady=5)
    button_update.grid(row=3, column=2, padx=5, pady=5)
    button_clear.grid(row=3, column=3, padx=5, pady=5)

    # Create the table view
    treeview = ttk.Treeview(content_frame, columns=("ID", "Nama Pembeli", "Tanggal Transaksi", "Tanggal Jatuh Tempo", "Pembayaran"), show="headings")
    treeview.heading("ID", text="No. Penjualan")
    treeview.heading("Nama Pembeli", text="Nama Pembeli")
    treeview.heading("Tanggal Transaksi", text="Tanggal Transaksi")
    treeview.heading("Tanggal Jatuh Tempo", text="Tanggal Jatuh Tempo")
    treeview.heading("Pembayaran", text="Pembayaran")

    treeview.column("ID", width=180)
    treeview.column("Nama Pembeli", width=180)
    treeview.column("Tanggal Transaksi", width=180)
    treeview.column("Tanggal Jatuh Tempo", width=180)
    treeview.column("Pembayaran", width=180)

    treeview.grid(row=4, column=0, columnspan=4, padx=1, pady=1)

    def on_select(event):
        selected_item = treeview.focus()
        if selected_item:
            penjualan_id, nama_pembeli, tanggal_transaksi, tanggal_tempo, pembayaran = treeview.item(selected_item, "values")
            entry_nama.delete(0, tk.END)
            entry_nama.insert(tk.END, nama_pembeli)
            entry_tanggal.delete(0, tk.END)
            entry_tanggal.insert(tk.END, tanggal_transaksi)
            entry_tanggal_tempo.delete(0, tk.END)
            entry_tanggal_tempo.insert(tk.END, tanggal_tempo)
            entry_pembayaran.delete(0, tk.END)
            entry_pembayaran.insert(tk.END, pembayaran)
            
            # Refresh the DetailJual table view based on the selected Penjualan
            refresh_detailjual_table(penjualan_id)

    treeview.bind("<<TreeviewSelect>>", on_select)


    # Function to add a new DetailJual record
    def add_detailjual():
        # Get the selected Penjualan ID
        selected_item = treeview.focus()
        if not selected_item:
            messagebox.showerror("Error", "No Penjualan record selected")
            return
        penjualan_id = treeview.item(selected_item, "values")[0]
        
        # Get the selected ID from the combobox
        selected_produk_id = produk_dropdown.get()
        c.execute("SELECT id_produk, nama_produk FROM Produk")
        produk_rows = c.fetchall()
        ## CARI STOK YANG ADA DI DETAILBELI BASED ON ID PRODUK
        for row in produk_rows:
            if selected_produk_id == row[1]:
                produk_id = row[0]
                
        # Get the input from the entry field
        harga_jual = entry_harga_jual.get()
        jumlah_jual = entry_jumlah.get()

        c.execute("SELECT SUM(jumlah_produk) FROM DetailBeli WHERE id_produk=?", (produk_id,))
        total_buy = c.fetchone()[0]

        # Calculate total quantity sold
        c.execute("SELECT SUM(jumlah) FROM DetailJual WHERE id_produk=?", (produk_id,))
        total_sold = c.fetchone()[0]

        # Calculate the remaining quantity in inventory
        if total_buy is None:
            total_buy = 0
        if total_sold is None:
            total_sold = 0

        remaining_quantity = total_buy - total_sold

        if int(jumlah_jual) > remaining_quantity:
            messagebox.showerror("Error", f"Stok tidak tersedia, stok hanya tersedia dengan jumlah : {remaining_quantity}")
            return

        
        # Insert the DetailJual data into the database
        c.execute("INSERT INTO DetailJual (id_penjualan, id_produk, harga_jual, jumlah) VALUES (?, ?, ?, ?)",
                (penjualan_id, produk_id, harga_jual, jumlah_jual))
        conn.commit()
        
        # Refresh the DetailJual table view
        refresh_detailjual_table(penjualan_id)


    # Function to update a DetailJual record
    def update_detailjual():
        selected_item = detailjual_treeview.focus()
        if selected_item:
            # Get the detailjual_id from the selected item
            detailjual_id = detailjual_treeview.item(selected_item)["values"][0]
            
            # Retrieve the modified values from the entry fields and dropdown
            # Get the selected ID from the combobox
            selected_produk_id = produk_dropdown.get()
            c.execute("SELECT id_produk, nama_produk FROM Produk")
            produk_rows = c.fetchall()
            for row in produk_rows:
                if selected_produk_id == row[1]:
                    updated_produk = row[0]
            
            updated_harga = entry_harga_jual.get()
            updated_jumlah = entry_jumlah.get()

            # Update the DetailJual record in the database
            c.execute("UPDATE DetailJual SET id_produk=?, harga_jual=?, jumlah=? WHERE id_detailjual=?", 
                    (updated_produk, updated_harga, updated_jumlah, detailjual_id))
            conn.commit()
            
            # Refresh the DetailJual table view
            penjualan_id = treeview.item(treeview.selection(), "values")[0]
            refresh_detailjual_table(penjualan_id)
            
    # Function to delete a DetailJual record
    def delete_detailjual():
        selected_item = detailjual_treeview.selection()
        if selected_item:
            # Get the detailjual_id from the selected item
            detailjual_id = detailjual_treeview.item(selected_item)["values"][0]
            
            # Delete the DetailJual record from the database
            c.execute("DELETE FROM DetailJual WHERE id_detailjual=?", (detailjual_id,))
            conn.commit()
            
            # Refresh the DetailJual table view
            penjualan_id = treeview.item(treeview.selection(), "values")[0]
            refresh_detailjual_table(penjualan_id)

    def refresh_detailjual_table(penjualan_id):
        # Clear existing table data
        for row in detailjual_treeview.get_children():
            detailjual_treeview.delete(row)
        
        # Fetch and display data from the database
        c.execute("SELECT * FROM DetailJual WHERE id_penjualan=?", (penjualan_id,))
        rows = c.fetchall()
        for row in rows:
            id_detailjual = row[0]
            c.execute("SELECT id_produk, nama_produk FROM Produk")
            produk_rows = c.fetchall()
            for rowss in produk_rows:
                if row[2] == rowss[0]:
                    nama_produk = rowss[1]
            harga_jual = row[3]
            jumlah_beli = row[4]
            
            # Insert data into the table view
            detailjual_treeview.insert("", tk.END, values=(id_detailjual, nama_produk, harga_jual, jumlah_beli))

    # Create the DetailJual table view
    detailjual_treeview = ttk.Treeview(content_frame, columns=("ID Detail Jual", "ID Produk", "Harga Jual", "Jumlah"), show="headings")
    detailjual_treeview.heading("ID Detail Jual", text="No. Detail Jual")
    detailjual_treeview.heading("ID Produk", text="Nama Produk")
    detailjual_treeview.heading("Harga Jual", text="Harga Jual")
    detailjual_treeview.heading("Jumlah", text="Jumlah")
    detailjual_treeview.column("ID Detail Jual", width=180)
    detailjual_treeview.column("ID Produk", width=180)
    detailjual_treeview.column("Harga Jual", width=180)
    detailjual_treeview.column("Jumlah", width=180)

    detailjual_treeview.grid(row=9, column=0, columnspan=4, padx=5, pady=5)

    def on_select_detailjual(event):
        selected_item = detailjual_treeview.focus()
        if selected_item:
            detailjual_id, nama_produk, harga_jual, jumlah_beli = detailjual_treeview.item(selected_item, "values")
            entry_jumlah.delete(0, tk.END)
            entry_jumlah.insert(tk.END, jumlah_beli)
            entry_harga_jual.delete(0, tk.END)
            entry_harga_jual.insert(tk.END, harga_jual)
            produk_dropdown.set(nama_produk)

    detailjual_treeview.bind("<<TreeviewSelect>>", on_select_detailjual)

    def clear_detailjual_fields():
        entry_jumlah.delete(0, tk.END)
        entry_harga_jual.delete(0, tk.END)
        produk_dropdown.set("")  # Clear the selection

    # Create the entry fields for DetailJual
    # Function to fetch and populate the Produk dropdown
    def populate_produk_dropdown():
        c.execute("SELECT id_produk, nama_produk FROM Produk")
        produk_rows = c.fetchall()

        listproduk = []
        listidprod = []
        for row in produk_rows:
            idprod = row[0]
            namaprod = row [1]
            listidprod.append(idprod)
            listproduk.append(namaprod)
        
        # produk_dropdown['values'] = listproduk
        return listproduk

    listproduk = populate_produk_dropdown()

    # Create the Produk dropdown
    produk_label = customtkinter.CTkLabel(content_frame, text="Produk", text_color="black")
    produk_dropdown = customtkinter.CTkComboBox(content_frame, values=listproduk)
    produk_label.grid(row=5, column=0, padx=5, pady=5)
    produk_dropdown.grid(row=5, column=1, padx=5, pady=5)

    label_harga_jual = customtkinter.CTkLabel(content_frame, text="Harga Jual", text_color="black")
    entry_harga_jual = customtkinter.CTkEntry(content_frame)

    label_harga_jual.grid(row=6, column=0, padx=5, pady=5)
    entry_harga_jual.grid(row=6, column=1, padx=5, pady=5)

    label_jumlah = customtkinter.CTkLabel(content_frame, text="Jumlah", text_color="black")
    entry_jumlah = customtkinter.CTkEntry(content_frame)

    label_jumlah.grid(row=7, column=0, padx=5, pady=5)
    entry_jumlah.grid(row=7, column=1, padx=5, pady=5)

    # Create the button for adding DetailJual
    button_add_detailjual = customtkinter.CTkButton(content_frame, text="Add", command=add_detailjual, width=100)
    button_add_detailjual.grid(row=8, column=0, padx=5, pady=5)

    # Create the buttons for updating and deleting DetailJual
    button_update_detailjual = customtkinter.CTkButton(content_frame, text="Update", command=update_detailjual, width=100)
    button_delete_detailjual = customtkinter.CTkButton(content_frame, text="Delete", command=delete_detailjual, width=100)
    button_clear_detailjual = customtkinter.CTkButton(content_frame, text="Clear", command=clear_detailjual_fields, width=100)
    button_update_detailjual.grid(row=8, column=2, padx=5, pady=5)
    button_delete_detailjual.grid(row=8, column=1, padx=5, pady=5)
    button_clear_detailjual.grid(row=8, column=3, padx=5, pady=5)

    # Fetch and display initial data in the table view
    refresh_table()

def show_transaksilain():
    # Hapus konten sebelumnya (jika ada)
    for widget in content_frame.winfo_children():
        widget.destroy()

    # Function to refresh the table view
    def refresh_table():
        # Clear existing table data
        for row in treeview.get_children():
            treeview.delete(row)
        
        # Fetch and display data from the database
        c.execute("SELECT * FROM Transaksi")
        rows = c.fetchall()
        for row in rows:
            transaksi_id = row[0]
            tanggal = row[1]
            penerima = row[2]
            jenis = row[3]
            keterangan = row[4]
            pembayaran = row[5]
            
            # Insert data into the table view
            treeview.insert("", tk.END, values=(transaksi_id, tanggal, penerima, jenis, keterangan, pembayaran))

    # Function to add a new penerimaan record
    def add_transaksi():
        # Fetch values from the entry fields
        tanggal = entry_tanggal.get_date().strftime('%Y-%m-%d')
        penerima = entry_penerima.get()
        jenis = entry_jenis.get()
        keterangan = entry_keterangan.get()
        pembayaran = entry_pembayaran.get()
        
        # Insert new penerimaan record into the database
        c.execute("INSERT INTO Transaksi (tanggal_transaksi, penerima, jenis, keterangan, pembayaran) VALUES (?, ?, ?, ?, ?)",
                (tanggal, penerima, jenis, keterangan, pembayaran))
        conn.commit()
        
        # Refresh the table view
        refresh_table()

    # Function to delete a penerimaan record
    def delete_transaksi():
        # Get the selected item from the table view
        selected_item = treeview.selection()
        if selected_item:
            # Get the penerimaan_id from the selected item
            transaksi_id = treeview.item(selected_item)["values"][0]
            
            # Delete the penerimaan record from the database
            c.execute("DELETE FROM Transaksi WHERE id_transaksi=?", (transaksi_id,))
            conn.commit()
            
            # Refresh the table view
            refresh_table()

    # Function to update a penerimaan record
    def update_transaksi():
        # Get the selected item from the table view
        selected_item = treeview.focus()
        if selected_item:
            # Get the penerimaan_id from the selected item
            transaksi_id = treeview.item(selected_item)["values"][0]

            # Perform database update based on modified values
            updated_tanggal = entry_tanggal.get_date().strftime('%Y-%m-%d')
            updated_penerima = entry_penerima.get()
            updated_jenis = entry_jenis.get()
            updated_keterangan = entry_keterangan.get()
            updated_pembayaran = entry_pembayaran.get()

            c.execute("UPDATE Penerimaan SET tanggal_transaksi=?, penerima=?, jenis=?, keterangan=?, pembayaran=? WHERE id_penerimaan=?", 
                    (updated_tanggal, updated_penerima, updated_jenis, updated_keterangan, updated_pembayaran, transaksi_id))

            # Commit the changes to the database
            conn.commit()

            # Refresh the table view
            refresh_table()

    # Function to clear the entry fields
    def clear_fields():
        entry_tanggal.set_date(None)
        entry_penerima.delete(0, tk.END)
        entry_jenis.set()
        entry_keterangan.delete(0, tk.END)
        entry_pembayaran.delete(0, tk.END)

    # Create the form widgets
    label_tanggal = customtkinter.CTkLabel(content_frame, text="Tanggal")
    entry_tanggal = DateEntry(content_frame, width=29, background='darkblue', foreground='white', date_pattern='dd/mm/yyyy')

    daftarjenis = ["Pengambilan Pribadi",
                   "Tambahan Modal",
                   "Pemberian Pinjaman",
                   "Penerimaan Pinjaman",
                   "Pengeluaran Kas Lainnya",
                   "Penerimaan Kas Lainnya",
                   ]
    
    label_penerima = customtkinter.CTkLabel(content_frame, text="Penerima/Pemberi")
    entry_penerima = customtkinter.CTkEntry(content_frame, width=200)

    label_jenis = customtkinter.CTkLabel(content_frame, text="Jenis")
    entry_jenis = customtkinter.CTkComboBox(content_frame, values=daftarjenis, width=200)

    label_keterangan = customtkinter.CTkLabel(content_frame, text="Keterangan")
    entry_keterangan = customtkinter.CTkEntry(content_frame, width=200)

    label_pembayaran = customtkinter.CTkLabel(content_frame, text="Pembayaran")
    entry_pembayaran = customtkinter.CTkEntry(content_frame, width=200)

    button_add = customtkinter.CTkButton(content_frame, text="Add", command=add_transaksi, width=100)
    button_delete = customtkinter.CTkButton(content_frame, text="Delete", command=delete_transaksi, width=100)
    button_update = customtkinter.CTkButton(content_frame, text="Update", command=update_transaksi, width=100)
    button_clear = customtkinter.CTkButton(content_frame, text="Clear", command=clear_fields, width=100)

    # Position the form widgets
    label_tanggal.grid(row=0, column=0, padx=5, pady=5)
    label_penerima.grid(row=1, column=0, padx=5, pady=5)
    label_jenis.grid(row=2, column=0, padx=5, pady=5)
    label_keterangan.grid(row=3, column=0, padx=5, pady=5)
    label_pembayaran.grid(row=4, column=0, padx=5, pady=5)

    entry_tanggal.grid(row=0, column=1, padx=5, pady=5)
    entry_penerima.grid(row=1, column=1, padx=5, pady=5)
    entry_jenis.grid(row=2, column=1, padx=5, pady=5)
    entry_keterangan.grid(row=3, column=1, padx=5, pady=5)
    entry_pembayaran.grid(row=4, column=1, padx=5, pady=5)

    button_add.grid(row=5, column=0, padx=1, pady=1)
    button_delete.grid(row=5, column=1, padx=1, pady=1)
    button_update.grid(row=5, column=2, padx=1, pady=1)
    button_clear.grid(row=5, column=3, padx=1, pady=1)

    # Create the table view
    treeview = ttk.Treeview(content_frame, columns=("ID", "Tanggal", "Penerima", "Jenis", "Keterangan", "Pembayaran"), show="headings", height=25)
    treeview.heading("ID", text="ID")
    treeview.heading("Tanggal", text="Tanggal")
    treeview.heading("Penerima", text="Penerima/Pemberi")
    treeview.heading("Jenis", text="Jenis")
    treeview.heading("Keterangan", text="Keterangan")
    treeview.heading("Pembayaran", text="Pembayaran")

    treeview.column("ID", width=50)
    treeview.column("Tanggal", width=100)
    treeview.column("Penerima", width=150)
    treeview.column("Jenis", width=150)
    treeview.column("Keterangan", width=200)
    treeview.column("Pembayaran", width=100)

    treeview.grid(row=6, columnspan=4, padx=5, pady=30)

    def on_select(event):
        selected_item = treeview.focus()
        if selected_item:
            penerimaan_id, tanggal, penerima, jenis, keterangan, pembayaran = treeview.item(selected_item, "values")
            entry_tanggal.delete(0, tk.END)
            entry_tanggal.insert(tk.END, tanggal)
            entry_penerima.delete(0, tk.END)
            entry_penerima.insert(tk.END, penerima)
            entry_jenis.set(jenis)
            entry_keterangan.delete(0, tk.END)
            entry_keterangan.insert(tk.END, keterangan)
            entry_pembayaran.delete(0, tk.END)
            entry_pembayaran.insert(tk.END, pembayaran)
            
    treeview.bind("<<TreeviewSelect>>", on_select)

    # Fetch and display initial data in the table view
    refresh_table()

def show_pembayaran():
    # Hapus konten sebelumnya (jika ada)
    for widget in content_frame.winfo_children():
        widget.destroy()

    # Function to refresh the table view
    def refresh_table():
        # Clear existing table data
        for row in treeview.get_children():
            treeview.delete(row)
        
        # Fetch and display data from the database
        c.execute("SELECT * FROM Pembayaran")
        rows = c.fetchall()
        for row in rows:
            pembayaran_id = row[0]
            tanggal = row[1]
            penerima = row[2]
            jenis = row[3]
            keterangan = row[4]
            pembayaran = row[5]
            
            # Insert data into the table view
            treeview.insert("", tk.END, values=(pembayaran_id, penerima, tanggal, jenis, keterangan, pembayaran))

    # Function to add a new pengeluaran record
    def add_pembayaran():
        # Fetch values from the entry fields
        tanggal = entry_tanggal.get_date().strftime('%Y-%m-%d')
        penerima = entry_jenis.get()
        jenis = entry_jenis.get()
        keterangan = entry_keterangan.get()
        pembayaran = entry_pembayaran.get()
        
        # Insert new pengeluaran record into the database
        c.execute("INSERT INTO Pembayaran (tanggal_transaksi, penerima, jenis, keterangan, pembayaran) VALUES (?, ?, ?, ?, ?)",
                (tanggal, penerima, jenis, keterangan, pembayaran))
        conn.commit()
        
        # Refresh the table view
        refresh_table()

    # Function to delete a pengeluaran record
    def delete_pembayaran():
        # Get the selected item from the table view
        selected_item = treeview.selection()
        if selected_item:
            # Get the pengeluaran_id from the selected item
            pembayaran_id = treeview.item(selected_item)["values"][0]
            
            # Delete the pengeluaran record from the database
            c.execute("DELETE FROM Pembayaran WHERE id_pembayaran=?", (pembayaran_id,))
            conn.commit()
            
            # Refresh the table view
            refresh_table()

    # Function to update a pengeluaran record
    def update_pembayaran():
        # Get the selected item from the table view
        selected_item = treeview.focus()
        if selected_item:
            # Get the pengeluaran_id from the selected item
            pembayaran_id = treeview.item(selected_item)["values"][0]

            # Perform database update based on modified values
            updated_tanggal = entry_tanggal.get_date().strftime('%Y-%m-%d')
            updated_penerima = entry_penerima.get()
            updated_jenis = entry_jenis.get()
            updated_keterangan = entry_keterangan.get()
            updated_pembayaran = entry_pembayaran.get()

            c.execute("UPDATE Pembayaran SET tanggal_transaksi=?, penerima=?, jenis=?, keterangan=?, pembayaran=? WHERE id_pembayaran=?", 
                    (updated_tanggal, updated_penerima, updated_jenis, updated_keterangan, updated_pembayaran, pembayaran_id))

            # Commit the changes to the database
            conn.commit()

            # Refresh the table view
            refresh_table()

    # Function to clear the entry fields
    def clear_fields():
        entry_tanggal.set_date(None)
        entry_penerima.delete(0, tk.END)
        entry_jenis.set()
        entry_keterangan.delete(0, tk.END)
        entry_pembayaran.delete(0, tk.END)

    # Create the form widgets
    label_tanggal = customtkinter.CTkLabel(content_frame, text="Tanggal")
    entry_tanggal = DateEntry(content_frame, width=29, background='darkblue', foreground='white', date_pattern='dd/mm/yyyy')

    daftarjenis = ["Pembayaran Gaji dan Upah",
                   "Pembayaran Listrik",
                   "Pembayaran Internet dan Telepon",
                   "Pembayaran Maintenance",
                   "Pembayaran Bahan Bakar",
                   ]
    
    label_penerima = customtkinter.CTkLabel(content_frame, text="Penerima")
    entry_penerima = customtkinter.CTkEntry(content_frame, width=200)

    label_jenis = customtkinter.CTkLabel(content_frame, text="Jenis")
    entry_jenis = customtkinter.CTkComboBox(content_frame,values=daftarjenis, width=200)

    label_keterangan = customtkinter.CTkLabel(content_frame, text="Keterangan")
    entry_keterangan = customtkinter.CTkEntry(content_frame, width=200)

    label_pembayaran = customtkinter.CTkLabel(content_frame, text="Pembayaran")
    entry_pembayaran = customtkinter.CTkEntry(content_frame, width=200)

    button_add = customtkinter.CTkButton(content_frame, text="Add", command=add_pembayaran, width=100)
    button_delete = customtkinter.CTkButton(content_frame, text="Delete", command=delete_pembayaran, width=100)
    button_update = customtkinter.CTkButton(content_frame, text="Update", command=update_pembayaran, width=100)
    button_clear = customtkinter.CTkButton(content_frame, text="Clear", command=clear_fields, width=100)

    # Position the form widgets
    label_tanggal.grid(row=0, column=0, padx=5, pady=5)
    label_penerima.grid(row=1, column=0, padx=5, pady=5)
    label_jenis.grid(row=2, column=0, padx=5, pady=5)
    label_keterangan.grid(row=3, column=0, padx=5, pady=5)
    label_pembayaran.grid(row=4, column=0, padx=5, pady=5)

    entry_tanggal.grid(row=0, column=1, padx=5, pady=5)
    entry_penerima.grid(row=1, column=1, padx=5, pady=5)
    entry_jenis.grid(row=2, column=1, padx=5, pady=5)
    entry_keterangan.grid(row=3, column=1, padx=5, pady=5)
    entry_pembayaran.grid(row=4, column=1, padx=5, pady=5)

    button_add.grid(row=5, column=0, padx=1, pady=1)
    button_delete.grid(row=5, column=1, padx=1, pady=1)
    button_update.grid(row=5, column=2, padx=1, pady=1)
    button_clear.grid(row=5, column=3, padx=1, pady=1)

    # Create the table view
    treeview = ttk.Treeview(content_frame, columns=("ID", "Tanggal", "Penerima", "Jenis", "Keterangan", "Pembayaran"), show="headings", height=25)
    treeview.heading("ID", text="ID")
    treeview.heading("Tanggal", text="Tanggal")
    treeview.heading("Penerima", text="Penerima")
    treeview.heading("Jenis", text="Jenis")
    treeview.heading("Keterangan", text="Keterangan")
    treeview.heading("Pembayaran", text="Pembayaran")

    treeview.column("ID", width=50)
    treeview.column("Tanggal", width=100)
    treeview.column("Penerima", width=200)
    treeview.column("Jenis", width=150)
    treeview.column("Keterangan", width=200)
    treeview.column("Pembayaran", width=100)

    treeview.grid(row=6, columnspan=4, padx=5, pady=30)

    def on_select(event):
        selected_item = treeview.focus()
        if selected_item:
            pembayaran_id, tanggal, penerima, jenis, keterangan, pembayaran = treeview.item(selected_item, "values")
            entry_tanggal.delete(0, tk.END)
            entry_tanggal.insert(tk.END, tanggal)
            entry_penerima.delete(0, tk.END)
            entry_penerima.insert(tk.END, penerima)
            entry_jenis.set(jenis)
            entry_keterangan.delete(0, tk.END)
            entry_keterangan.insert(tk.END, keterangan)
            entry_pembayaran.delete(0, tk.END)
            entry_pembayaran.insert(tk.END, pembayaran)
            
    treeview.bind("<<TreeviewSelect>>", on_select)

    # Fetch and display initial data in the table view
    refresh_table()

def show_piutang():
    # Hapus konten sebelumnya (jika ada)
    for widget in content_frame.winfo_children():
        widget.destroy()

    # Function to refresh the table view
    def refresh_table():
        # Clear existing table data
        for row in treeview.get_children():
            treeview.delete(row)

        # Fetch and display data from the database
        c.execute("SELECT * FROM Piutang")
        rows = c.fetchall()
        for row in rows:
            piutang_id = row[0]
            penjualan_id = row[1]
            tanggal_bayar = row[2]
            pembayaran = row[3]

            # # Get the nama_pembeli from the Penjualan table based on penjualan_id
            # c.execute("SELECT nama_pembeli FROM Penjualan WHERE id_penjualan=?", (penjualan_id,))
            # result = c.fetchone()
            # if result:
            #     nama_pembeli = result[0]
            # else:
            #     nama_pembeli = ""

            # Insert data into the table view
            treeview.insert("", tk.END, values=(piutang_id, penjualan_id, tanggal_bayar, pembayaran))

    # Function to fetch penjualan IDs
    def fetch_penjualan_ids():
        c.execute("SELECT id_penjualan FROM Penjualan")
        penjualan_ids = c.fetchall()
        return [str(id[0]) for id in penjualan_ids]

    # Function to add a new piutang record
    def add_piutang():
        # Fetch values from the entry fields
        penjualan_id = combo_penjualan.get()
        tanggal_bayar = entry_tanggal_bayar.get()
        pembayaran = entry_pembayaran.get()

        # Insert new piutang record into the database
        c.execute("INSERT INTO Piutang (id_penjualan, tanggal_bayar, pembayaran) VALUES (?, ?, ?)",
                (penjualan_id, tanggal_bayar, pembayaran))
        conn.commit()

        # Refresh the table view
        refresh_table()

    # Function to delete a piutang record
    def delete_piutang():
        # Get the selected item from the table view
        selected_item = treeview.selection()
        if selected_item:
            # Get the piutang_id from the selected item
            piutang_id = treeview.item(selected_item)["values"][0]

            # Delete the piutang record from the database
            c.execute("DELETE FROM Piutang WHERE id_piutang=?", (piutang_id,))
            conn.commit()

            # Refresh the table view
            refresh_table()

    # Function to update a piutang record
    def update_piutang():
        # Get the selected item from the table view
        selected_item = treeview.focus()
        if selected_item:
            # Get the piutang_id from the selected item
            piutang_id = treeview.item(selected_item)["values"][0]

            # Perform database update based on modified values
            updated_penjualan_id = combo_penjualan.get()
            updated_tanggal_bayar = entry_tanggal_bayar.get()
            updated_pembayaran = entry_pembayaran.get()

            c.execute("UPDATE Piutang SET id_penjualan=?, tanggal_bayar=?, pembayaran=? WHERE id_piutang=?",
                    (updated_penjualan_id, updated_tanggal_bayar, updated_pembayaran, piutang_id))

            # Commit the changes to the database
            conn.commit()

            # Refresh the table view
            refresh_table()

    # Function to clear the entry fields
    def clear_fields():
        combo_penjualan.current(0)
        entry_tanggal_bayar.delete(0, tk.END)
        entry_pembayaran.delete(0, tk.END)

    # Create the form widgets
    label_penjualan = customtkinter.CTkLabel(content_frame, text="Penjualan")
    combo_penjualan = customtkinter.CTkComboBox(content_frame, values=fetch_penjualan_ids(), width=200)

    label_tanggal_bayar = customtkinter.CTkLabel(content_frame, text="Tanggal Bayar")
    entry_tanggal_bayar = DateEntry(content_frame, width=29, background='darkblue', foreground='white', date_pattern='yyyy-mm-dd')

    label_pembayaran = customtkinter.CTkLabel(content_frame, text="Pembayaran")
    entry_pembayaran = customtkinter.CTkEntry(content_frame, width=200)

    button_add = customtkinter.CTkButton(content_frame, text="Add", command=add_piutang, width=100)
    button_delete = customtkinter.CTkButton(content_frame, text="Delete", command=delete_piutang, width=100)
    button_update = customtkinter.CTkButton(content_frame, text="Update", command=update_piutang, width=100)
    button_clear = customtkinter.CTkButton(content_frame, text="Clear", command=clear_fields, width=100)

    # Position the form widgets
    label_penjualan.grid(row=0, column=0, padx=5, pady=5)
    combo_penjualan.grid(row=0, column=1, padx=5, pady=5)

    label_tanggal_bayar.grid(row=1, column=0, padx=5, pady=5)
    entry_tanggal_bayar.grid(row=1, column=1, padx=5, pady=5)

    label_pembayaran.grid(row=2, column=0, padx=5, pady=5)
    entry_pembayaran.grid(row=2, column=1, padx=5, pady=5)

    button_add.grid(row=3, column=0, padx=1, pady=1)
    button_delete.grid(row=3, column=1, padx=1, pady=1)
    button_update.grid(row=3, column=2, padx=1, pady=1)
    button_clear.grid(row=3, column=3, padx=1, pady=1)

    # Create the table view
    treeview = ttk.Treeview(content_frame, columns=("ID", "ID Penjualan", "Tanggal Bayar", "Pembayaran"), show="headings", height=30)
    treeview.heading("ID", text="ID")
    treeview.heading("ID Penjualan", text="ID Penjualan")
    treeview.heading("Tanggal Bayar", text="Tanggal Bayar")
    treeview.heading("Pembayaran", text="Pembayaran")

    treeview.column("ID", width=50)
    treeview.column("ID Penjualan", width=100)
    treeview.column("Tanggal Bayar", width=120)
    treeview.column("Pembayaran", width=100)

    treeview.grid(row=4, columnspan=3, padx=5, pady=30)

    def on_select(event):
        selected_item = treeview.focus()
        if selected_item:
            piutang_id, penjualan_id, tanggal_bayar, pembayaran = treeview.item(selected_item, "values")
            combo_penjualan.set(penjualan_id)
            entry_tanggal_bayar.delete(0, tk.END)
            entry_tanggal_bayar.insert(tk.END, tanggal_bayar)
            entry_pembayaran.delete(0, tk.END)
            entry_pembayaran.insert(tk.END, pembayaran)

    treeview.bind("<<TreeviewSelect>>", on_select)

    # Fetch and display initial data in the table view
    refresh_table()

def show_utang():
    # Hapus konten sebelumnya (jika ada)
    for widget in content_frame.winfo_children():
        widget.destroy()

    # Function to refresh the table view
    def refresh_table():
        # Clear existing table data
        for row in treeview.get_children():
            treeview.delete(row)
        
        # Fetch and display data from the database
        c.execute("SELECT * FROM Utang")
        rows = c.fetchall()
        for row in rows:
            utang_id = row[0]
            pembelian_id = row[1]
            tanggal_bayar = row[2]
            pembayaran = row[3]
            
            # Insert data into the table view
            treeview.insert("", tk.END, values=(utang_id, pembelian_id, tanggal_bayar, pembayaran))

    # Function to fetch pembelian IDs
    def fetch_pembelian_ids():
        c.execute("SELECT id_pembelian FROM Pembelian")
        pembelian_ids = c.fetchall()
        return [str(id[0]) for id in pembelian_ids]

    # Function to add a new utang record
    def add_utang():
        pembelian_id = entry_pembelian_id.get()
        tanggal_bayar = entry_tanggal_bayar.get()
        pembayaran = entry_pembayaran.get()
        
        c.execute("INSERT INTO Utang (id_pembelian, tanggal_bayar, pembayaran) VALUES (?, ?, ?)",
                (pembelian_id, tanggal_bayar, pembayaran))
        conn.commit()
        
        refresh_table()

    # Function to delete an utang record
    def delete_utang():
        selected_item = treeview.selection()
        if selected_item:
            utang_id = treeview.item(selected_item)["values"][0]
            
            c.execute("DELETE FROM Utang WHERE id_utang=?", (utang_id,))
            conn.commit()
            
            refresh_table()

    # Function to update an utang record
    def update_utang():
        selected_item = treeview.focus()
        if selected_item:
            utang_id = treeview.item(selected_item)["values"][0]

            updated_pembelian_id = entry_pembelian_id.get()
            updated_tanggal_bayar = entry_tanggal_bayar.get()
            updated_pembayaran = entry_pembayaran.get()

            c.execute("UPDATE Utang SET id_pembelian=?, tanggal_bayar=?, pembayaran=? WHERE id_utang=?", 
                    (updated_pembelian_id, updated_tanggal_bayar, updated_pembayaran, utang_id))

            conn.commit()

            refresh_table()

    # Function to clear the entry fields
    def clear_fields():
        entry_pembelian_id.set('')
        entry_tanggal_bayar.delete(0, tk.END)
        entry_pembayaran.delete(0, tk.END)

    # Create the form widgets
    label_pembelian_id = customtkinter.CTkLabel(content_frame, text="ID Pembelian")
    entry_pembelian_id = customtkinter.CTkComboBox(content_frame, values=fetch_pembelian_ids(), width=200)

    label_tanggal_bayar = customtkinter.CTkLabel(content_frame, text="Tanggal Bayar")
    entry_tanggal_bayar = DateEntry(content_frame, width=29, background='darkblue', foreground='white', date_pattern='yyyy-mm-dd')

    label_pembayaran = customtkinter.CTkLabel(content_frame, text="Pembayaran")
    entry_pembayaran = customtkinter.CTkEntry(content_frame, width=200)

    button_add = customtkinter.CTkButton(content_frame, text="Add", command=add_utang)
    button_delete = customtkinter.CTkButton(content_frame, text="Delete", command=delete_utang)
    button_update = customtkinter.CTkButton(content_frame, text="Update", command=update_utang)
    button_clear = customtkinter.CTkButton(content_frame, text="Clear", command=clear_fields)

    # Position the form widgets
    label_pembelian_id.grid(row=0, column=0, padx=5, pady=5)
    entry_pembelian_id.grid(row=0, column=1, padx=5, pady=5)

    label_tanggal_bayar.grid(row=1, column=0, padx=5, pady=5)
    entry_tanggal_bayar.grid(row=1, column=1, padx=5, pady=5)

    label_pembayaran.grid(row=2, column=0, padx=5, pady=5)
    entry_pembayaran.grid(row=2, column=1, padx=5, pady=5)

    button_add.grid(row=3, column=0, padx=1, pady=1)
    button_delete.grid(row=3, column=1, padx=1, pady=1)
    button_update.grid(row=3, column=2, padx=1, pady=1)
    button_clear.grid(row=3, column=3, padx=1, pady=1)

    # Create the table view
    treeview = ttk.Treeview(content_frame, columns=("ID", "ID Pembelian", "Tanggal Bayar", "Pembayaran"), show="headings", height=30)
    treeview.heading("ID", text="ID")
    treeview.heading("ID Pembelian", text="ID Pembelian")
    treeview.heading("Tanggal Bayar", text="Tanggal Bayar")
    treeview.heading("Pembayaran", text="Pembayaran")

    treeview.column("ID", width=50)
    treeview.column("ID Pembelian", width=100)
    treeview.column("Tanggal Bayar", width=120)
    treeview.column("Pembayaran", width=100)

    treeview.grid(row=4, columnspan=3, padx=5, pady=30)
    def on_select(event):
        selected_item = treeview.focus()
        if selected_item:
            utang_id, pembelian_id, tanggal_bayar, pembayaran = treeview.item(selected_item, "values")
            entry_pembelian_id.set(pembelian_id)
            entry_tanggal_bayar.delete(0, tk.END)
            entry_tanggal_bayar.insert(tk.END, tanggal_bayar)
            entry_pembayaran.delete(0, tk.END)
            entry_pembayaran.insert(tk.END, pembayaran)
            
    treeview.bind("<<TreeviewSelect>>", on_select)

    # Fetch and display initial data in the table view
    refresh_table()

def show_produk():
    # Hapus konten sebelumnya (jika ada)
    for widget in content_frame.winfo_children():
        widget.destroy()

        # Function to refresh the table view
    def refresh_table():
        # Clear existing table data
        for row in treeview.get_children():
            treeview.delete(row)

        # Fetch and display data from the database
        c.execute("SELECT * FROM Produk")
        rows = c.fetchall()
        for row in rows:
            produk_id = row[0]
            nama_produk = row[1]
            jenis = row[2]
            satuan = row[3]

            # Insert data into the table view
            treeview.insert("", tk.END, values=(produk_id, nama_produk, jenis, satuan))

    # Function to add a new produk record
    def add_produk():
        # Fetch values from the entry fields
        nama_produk = entry_nama.get()
        satuan = entry_satuan.get()
        jenis = entry_jenis.get()

        # Insert new produk record into the database
        c.execute("INSERT INTO Produk (nama_produk, jenis, satuan) VALUES (?, ?, ?)",
                (nama_produk, jenis, satuan))
        conn.commit()

        # Refresh the table view
        refresh_table()

    # Function to delete a produk record
    def delete_produk():
        # Get the selected item from the table view
        selected_item = treeview.selection()
        if selected_item:
            # Get the produk_id from the selected item
            produk_id = treeview.item(selected_item)["values"][0]

            # Delete the produk record from the database
            c.execute("DELETE FROM Produk WHERE id_produk=?", (produk_id,))
            conn.commit()

            # Refresh the table view
            refresh_table()

    # Function to update a produk record
    def update_produk():
        # Get the selected item from the table view
        selected_item = treeview.focus()
        if selected_item:
            # Get the produk_id from the selected item
            produk_id = treeview.item(selected_item)["values"][0]

            # Perform database update based on modified values
            updated_nama = entry_nama.get()
            updated_jenis = entry_jenis.get()
            updated_satuan = entry_satuan.get()

            c.execute("UPDATE Produk SET nama_produk=?, jenis=?, satuan=? WHERE id_produk=?",
                    (updated_nama, updated_jenis, updated_satuan, produk_id))

            # Commit the changes to the database
            conn.commit()

            # Refresh the table view
            refresh_table()

    # Function to clear the entry fields
    def clear_fields():
        entry_nama.delete(0, tk.END)
        entry_satuan.delete(0, tk.END)
        entry_jenis.delete(0, tk.END)

    # Create the form widgets
    label_nama = customtkinter.CTkLabel(content_frame, text="Nama Produk")
    entry_nama = customtkinter.CTkEntry(content_frame)

    label_jenis = customtkinter.CTkLabel(content_frame, text="Jenis")
    entry_jenis = customtkinter.CTkEntry(content_frame)

    label_satuan = customtkinter.CTkLabel(content_frame, text="Satuan")
    entry_satuan = customtkinter.CTkEntry(content_frame)

    button_add = customtkinter.CTkButton(content_frame, text="Add", command=add_produk)
    button_delete = customtkinter.CTkButton(content_frame, text="Delete", command=delete_produk)
    button_update = customtkinter.CTkButton(content_frame, text="Update", command=update_produk)
    button_clear = customtkinter.CTkButton(content_frame, text="Clear", command=clear_fields)

    # Position the form widgets
    label_nama.grid(row=0, column=0, padx=5, pady=5)
    entry_nama.grid(row=0, column=1, padx=5, pady=5)

    label_jenis.grid(row=1, column=0, padx=5, pady=5)
    entry_jenis.grid(row=1, column=1, padx=5, pady=5)

    label_satuan.grid(row=2, column=0, padx=5, pady=5)
    entry_satuan.grid(row=2, column=1, padx=5, pady=5)

    button_add.grid(row=3, column=0, padx=1, pady=1)
    button_delete.grid(row=3, column=1, padx=1, pady=1)
    button_update.grid(row=3, column=2, padx=1, pady=1)
    button_clear.grid(row=3, column=3, padx=1, pady=1)

    # Create the table view
    treeview = ttk.Treeview(content_frame, columns=("ID", "Nama Produk", "Jenis", "Satuan"), show="headings", height=30)
    treeview.heading("ID", text="ID")
    treeview.heading("Nama Produk", text="Nama Produk")
    treeview.heading("Jenis", text="Jenis")
    treeview.heading("Satuan", text="Satuan")

    treeview.column("ID", width=50)
    treeview.column("Nama Produk", width=150)
    treeview.column("Jenis", width=100)
    treeview.column("Satuan", width=100)

    treeview.grid(row=4, columnspan=3, padx=5, pady=30)

    def on_select(event):
        selected_item = treeview.focus()
        if selected_item:
            produk_id, nama_produk, jenis, satuan = treeview.item(selected_item, "values")
            entry_nama.delete(0, tk.END)
            entry_nama.insert(tk.END, nama_produk)
            entry_jenis.delete(0, tk.END)
            entry_jenis.insert(tk.END, jenis)
            entry_satuan.delete(0, tk.END)
            entry_satuan.insert(tk.END, satuan)

    treeview.bind("<<TreeviewSelect>>", on_select)

    # Fetch and display initial data in the table view
    refresh_table()

def show_supplier():
    # Hapus konten sebelumnya (jika ada)
    for widget in content_frame.winfo_children():
        widget.destroy()

    # Function to refresh the table view
    def refresh_table():
        # Clear existing table data
        for row in treeview.get_children():
            treeview.delete(row)
        
        # Fetch and display data from the database
        c.execute("SELECT * FROM Supplier")
        rows = c.fetchall()
        for row in rows:
            supplier_id = row[0]
            nama_supplier = row[1]
            alamat = row[2]
            nomor_hp = row[3]
            
            # Insert data into the table view
            treeview.insert("", tk.END, values=(supplier_id, nama_supplier, alamat, nomor_hp))

    # Function to add a new supplier record
    def add_supplier():
        # Fetch values from the entry fields
        nama_supplier = entry_nama.get()
        alamat = entry_alamat.get()
        nomor_hp = entry_nomor_hp.get()
        
        # Insert new supplier record into the database
        c.execute("INSERT INTO Supplier (nama_supplier, alamat, nomor_hp) VALUES (?, ?, ?)",
                (nama_supplier, alamat, nomor_hp))
        conn.commit()
        
        # Refresh the table view
        refresh_table()

    # Function to delete a supplier record
    def delete_supplier():
        # Get the selected item from the table view
        selected_item = treeview.selection()
        if selected_item:
            # Get the supplier_id from the selected item
            supplier_id = treeview.item(selected_item)["values"][0]
            
            # Delete the supplier record from the database
            c.execute("DELETE FROM Supplier WHERE id_supplier=?", (supplier_id,))
            conn.commit()
            
            # Refresh the table view
            refresh_table()

    # Function to update a supplier record
    def update_supplier():
        # Get the selected item from the table view
        selected_item = treeview.focus()
        if selected_item:
            # Get the supplier_id from the selected item
            supplier_id = treeview.item(selected_item)["values"][0]

            # Perform database update based on modified values
            updated_nama = entry_nama.get()
            updated_alamat = entry_alamat.get()
            updated_nomor_hp = entry_nomor_hp.get()

            c.execute("UPDATE Supplier SET nama_supplier=?, alamat=?, nomor_hp=? WHERE id_supplier=?", 
                    (updated_nama, updated_alamat, updated_nomor_hp, supplier_id))

            # Commit the changes to the database
            conn.commit()

            # Refresh the table view
            refresh_table()

    # Function to clear the entry fields
    def clear_fields():
        entry_nama.delete(0, tk.END)
        entry_alamat.delete(0, tk.END)
        entry_nomor_hp.delete(0, tk.END)

    # Create the form widgets
    label_nama = customtkinter.CTkLabel(content_frame, text="Nama Supplier")
    entry_nama = customtkinter.CTkEntry(content_frame)

    label_alamat = customtkinter.CTkLabel(content_frame, text="Alamat")
    entry_alamat = customtkinter.CTkEntry(content_frame)

    label_nomor_hp = customtkinter.CTkLabel(content_frame, text="Nomor HP")
    entry_nomor_hp = customtkinter.CTkEntry(content_frame)

    button_add = customtkinter.CTkButton(content_frame, text="Add", command=add_supplier)
    button_delete = customtkinter.CTkButton(content_frame, text="Delete", command=delete_supplier)
    button_update = customtkinter.CTkButton(content_frame, text="Update", command=update_supplier)
    button_clear = customtkinter.CTkButton(content_frame, text="Clear", command=clear_fields)

    # Position the form widgets
    label_nama.grid(row=0, column=0, padx=5, pady=5)
    entry_nama.grid(row=0, column=1, padx=5, pady=5)

    label_alamat.grid(row=1, column=0, padx=5, pady=5)
    entry_alamat.grid(row=1, column=1, padx=5, pady=5)

    label_nomor_hp.grid(row=2, column=0, padx=5, pady=5)
    entry_nomor_hp.grid(row=2, column=1, padx=5, pady=5)

    button_add.grid(row=3, column=0, padx=1, pady=1)
    button_delete.grid(row=3, column=1, padx=1, pady=1)
    button_update.grid(row=3, column=2, padx=1, pady=1)
    button_clear.grid(row=3, column=3, padx=1, pady=1)

    # Create the table view
    treeview = ttk.Treeview(content_frame, columns=("ID", "Nama Supplier", "Alamat", "Nomor HP"), show="headings", height=28)
    treeview.heading("ID", text="ID")
    treeview.heading("Nama Supplier", text="Nama Supplier")
    treeview.heading("Alamat", text="Alamat")
    treeview.heading("Nomor HP", text="Nomor HP")

    treeview.column("ID", width=50)
    treeview.column("Nama Supplier", width=150)
    treeview.column("Alamat", width=200)
    treeview.column("Nomor HP", width=100)

    treeview.grid(row=4, columnspan=3, padx=5, pady=30)

    def on_select(event):
        selected_item = treeview.focus()
        if selected_item:
            supplier_id, nama_supplier, alamat, nomor_hp = treeview.item(selected_item, "values")
            entry_nama.delete(0, tk.END)
            entry_nama.insert(tk.END, nama_supplier)
            entry_alamat.delete(0, tk.END)
            entry_alamat.insert(tk.END, alamat)
            entry_nomor_hp.delete(0, tk.END)
            entry_nomor_hp.insert(tk.END, nomor_hp)

    treeview.bind("<<TreeviewSelect>>", on_select)

    # Fetch and display initial data in the table view
    refresh_table()

def show_laporanpenjualan():
    # Hapus konten sebelumnya (jika ada)
    for widget in content_frame.winfo_children():
        widget.destroy()

    # Function to extract and display data based on date range
    def show_data():
        # Clear existing data in the table view
        for row in treeview.get_children():
            treeview.delete(row)
        
        # Fetch data from the database based on date range
        start_date = entry_start_date.get()
        end_date = entry_end_date.get()
        
        c.execute('''SELECT Produk.jenis, Produk.nama_produk, SUM(DetailJual.jumlah), Produk.satuan, SUM(DetailJual.harga_jual * DetailJual.jumlah) AS nilai_penjualan
                 FROM Produk
                 JOIN DetailJual ON Produk.id_produk = DetailJual.id_produk
                 JOIN Penjualan ON DetailJual.id_penjualan = Penjualan.id_penjualan
                 WHERE Penjualan.tanggal_jual BETWEEN ? AND ?
                 GROUP BY Produk.nama_produk''',
              (start_date, end_date))
        
        rows = c.fetchall()
        
        # Insert data into the table view
        for row in rows:
            jenis = row[0]
            produk = row[1]
            jumlah = row[2]
            satuan = row[3]
            nilai_penjualan = row[4]
            
            treeview.insert("", tk.END, values=(jenis, produk, jumlah, satuan, nilai_penjualan))

    # Create the form widgets
    label_start_date = customtkinter.CTkLabel(content_frame, text="Start Date")
    entry_start_date = DateEntry(content_frame, width=12, background='darkblue', foreground='white', date_pattern='yyyy-mm-dd')

    label_end_date = customtkinter.CTkLabel(content_frame, text="End Date")
    entry_end_date = DateEntry(content_frame, width=12, background='darkblue', foreground='white', date_pattern='yyyy-mm-dd')

    button_show_data = customtkinter.CTkButton(content_frame, text="Show Data", command=show_data)

    # Position the form widgets
    label_start_date.grid(row=0, column=0, padx=5, pady=5)
    entry_start_date.grid(row=0, column=1, padx=5, pady=5)

    label_end_date.grid(row=1, column=0, padx=5, pady=5)
    entry_end_date.grid(row=1, column=1, padx=5, pady=5)

    button_show_data.grid(row=2, column=0, columnspan=2, padx=5, pady=5)

    # Create the table view
    treeview = ttk.Treeview(content_frame, columns=("Jenis", "Produk", "Jumlah", "Satuan", "Nilai Penjualan"), show="headings", height=30)
    treeview.heading("Jenis", text="Jenis")
    treeview.heading("Produk", text="Produk")
    treeview.heading("Jumlah", text="Jumlah")
    treeview.heading("Satuan", text="Satuan")
    treeview.heading("Nilai Penjualan", text="Nilai Penjualan")

    treeview.column("Jenis", width=80)
    treeview.column("Produk", width=150)
    treeview.column("Jumlah", width=80)
    treeview.column("Satuan", width=80)
    treeview.column("Nilai Penjualan", width=120)

    treeview.grid(row=3, column=0, columnspan=2, padx=5, pady=20)

def show_laporanpembelian():
    # Hapus konten sebelumnya (jika ada)
    for widget in content_frame.winfo_children():
        widget.destroy()

    # Function to extract and display data based on date range
    def show_data():
        # Clear existing data in the table view
        for row in treeview.get_children():
            treeview.delete(row)
        
        # Fetch data from the database based on date range
        start_date = entry_start_date.get()
        end_date = entry_end_date.get()
        
        c.execute('''SELECT Produk.jenis, Produk.nama_produk, SUM(DetailBeli.jumlah_produk), Produk.satuan, SUM(DetailBeli.harga_beli * DetailBeli.jumlah_produk) AS nilai_pembelian
                 FROM Produk
                 JOIN DetailBeli ON Produk.id_produk = DetailBeli.id_produk
                 JOIN Pembelian ON DetailBeli.id_pembelian = Pembelian.id_pembelian
                 WHERE Pembelian.tanggal_beli BETWEEN ? AND ?
                 GROUP BY Produk.nama_produk''',
              (start_date, end_date))
        
        rows = c.fetchall()
        
        # Insert data into the table view
        for row in rows:
            jenis = row[0]
            produk = row[1]
            jumlah = row[2]
            satuan = row[3]
            nilai_pembelian = row[4]
            treeview.insert("", tk.END, values=(jenis, produk, jumlah, satuan, nilai_pembelian))

    # Create the form widgets
    label_start_date = customtkinter.CTkLabel(content_frame, text="Start Date")
    entry_start_date = DateEntry(content_frame, width=12, background='darkblue', foreground='white', date_pattern='yyyy-mm-dd')

    label_end_date = customtkinter.CTkLabel(content_frame, text="End Date")
    entry_end_date = DateEntry(content_frame, width=12, background='darkblue', foreground='white', date_pattern='yyyy-mm-dd')

    button_show_data = customtkinter.CTkButton(content_frame, text="Show Data", command=show_data)

    # Position the form widgets
    label_start_date.grid(row=0, column=0, padx=5, pady=5)
    entry_start_date.grid(row=0, column=1, padx=5, pady=5)

    label_end_date.grid(row=1, column=0, padx=5, pady=5)
    entry_end_date.grid(row=1, column=1, padx=5, pady=5)

    button_show_data.grid(row=2, column=0, columnspan=2, padx=5, pady=5)

    # Create the table view
    treeview = ttk.Treeview(content_frame, columns=("Jenis", "Produk", "Jumlah", "Satuan", "Nilai Pembelian"), show="headings", height=30)
    treeview.heading("Jenis", text="Jenis")
    treeview.heading("Produk", text="Produk")
    treeview.heading("Jumlah", text="Jumlah")
    treeview.heading("Satuan", text="Satuan")
    treeview.heading("Nilai Pembelian", text="Nilai Pembelian")

    treeview.column("Jenis", width=80)
    treeview.column("Produk", width=150)
    treeview.column("Jumlah", width=80)
    treeview.column("Satuan", width=80)
    treeview.column("Nilai Pembelian", width=120)

    treeview.grid(row=3, column=0, columnspan=2, padx=5, pady=20)

def show_persediaanproduk():
    # Hapus konten sebelumnya (jika ada)
    for widget in content_frame.winfo_children():
        widget.destroy()
        
    def calculate_products_inventory():
        c.execute("SELECT id_produk, nama_produk, jenis, satuan FROM Produk")
        produk_rows = c.fetchall()

        inventory = {}

        for row in produk_rows:
            produk_id = row[0]
            nama_produk = row[1]
            jenis = row[2]
            satuan = row[3]

            # Calculate total quantity produced
            c.execute("SELECT SUM(jumlah_produk) FROM DetailBeli WHERE id_produk=?", (produk_id,))
            total_buy = c.fetchone()[0]

            # Calculate total quantity sold
            c.execute("SELECT SUM(jumlah) FROM DetailJual WHERE id_produk=?", (produk_id,))
            total_sold = c.fetchone()[0]

            # Calculate the remaining quantity in inventory
            if total_buy is None:
                total_buy = 0
            if total_sold is None:
                total_sold = 0

            remaining_quantity = total_buy - total_sold

            # Add the product and its inventory details to the dictionary
            inventory[produk_id] = {
                'nama_produk': nama_produk,
                'jenis': jenis,
                'satuan': satuan,
                'remaining_quantity': remaining_quantity
            }

        if remaining_quantity <=0 :
            messagebox.showerror("Error", f"Sisa persediaan {nama_produk} : 0, silahkan lakukan restock!")
        
        return inventory


    # Generate and display the inventory report
    def generate_inventory_report():
        
        products_inventory = calculate_products_inventory()

        for produk_id, details in products_inventory.items():
            nama_produk = details['nama_produk']
            jenis = details['jenis']
            satuan = details['satuan']
            remaining_quantity = details['remaining_quantity']

            treeview.insert("", tk.END, values=(nama_produk, jenis, remaining_quantity,  satuan))
        

    treeview = ttk.Treeview(content_frame, columns=("Item", "Jenis", "Value",  "Satuan"), show="headings", height=20)
    treeview.heading("Item", text="Produk")
    treeview.heading("Jenis", text="Jenis")
    treeview.heading("Value", text="Jumlah")
    treeview.heading("Satuan", text="Satuan")

    treeview.column("Item", width=200)
    treeview.column("Jenis", width=150)
    treeview.column("Value", width=150)
    treeview.column("Satuan", width=150)

    treeview.grid(row=3, columnspan=2, padx=5, pady=5,)

    # Generate and display the inventory report
    generate_inventory_report()

def show_laporanutangpiutang():

    # Hapus konten sebelumnya (jika ada)
    for widget in content_frame.winfo_children():
        widget.destroy()
    
    # Create the table view
    treeviewutang = ttk.Treeview(content_frame, columns=("ID Pembelian", "Tanggal", "Nama Supplier", "Nilai Utang", "Pembayaran", "Sisa Utang"), show="headings", height=15)
    treeviewutang.heading("ID Pembelian", text="ID Pembelian")
    treeviewutang.heading("Tanggal", text="Tanggal")
    treeviewutang.heading("Nama Supplier", text="Nama Supplier")
    treeviewutang.heading("Nilai Utang", text="Nilai Utang")
    treeviewutang.heading("Pembayaran", text="Pembayaran")
    treeviewutang.heading("Sisa Utang", text="Sisa Utang")

    treeviewutang.column("ID Pembelian", width=100)
    treeviewutang.column("Tanggal", width=100)
    treeviewutang.column("Nama Supplier", width=150)
    treeviewutang.column("Nilai Utang", width=130)
    treeviewutang.column("Pembayaran", width=130)
    treeviewutang.column("Sisa Utang", width=130)

    treeviewutang.grid(row=0, column=0, padx=5, pady=30)

    # Create the table view
    treeviewpiutang = ttk.Treeview(content_frame, columns=("ID Penjualan", "Tanggal", "Nama Pembeli", "Nilai Piutang", "Pembayaran", "Sisa Piutang"), show="headings", height=15)
    treeviewpiutang.heading("ID Penjualan", text="ID Pembelian")
    treeviewpiutang.heading("Tanggal", text="Tanggal")
    treeviewpiutang.heading("Nama Pembeli", text="Nama Pembeli")
    treeviewpiutang.heading("Nilai Piutang", text="Nilai Piutang")
    treeviewpiutang.heading("Pembayaran", text="Pembayaran")
    treeviewpiutang.heading("Sisa Piutang", text="Sisa Piutang")

    treeviewpiutang.column("ID Penjualan", width=100)
    treeviewpiutang.column("Tanggal", width=100)
    treeviewpiutang.column("Nama Pembeli", width=150)
    treeviewpiutang.column("Nilai Piutang", width=130)
    treeviewpiutang.column("Pembayaran", width=130)
    treeviewpiutang.column("Sisa Piutang", width=130)

    treeviewpiutang.grid(row=1, column=0, padx=5, pady=30)

    # Function to refresh the table view
    def generate_utang_report():
        # Clear existing table data
        for row in treeviewutang.get_children():
            treeviewutang.delete(row)
    
        c.execute('''SELECT Pembelian.id_pembelian, Pembelian.tanggal_beli, Supplier.nama_supplier,
                        DetailBeli.harga_beli * DetailBeli.jumlah_produk - Pembelian.pembayaran AS total_utang, 
                        SUM(Utang.pembayaran)
                FROM Pembelian
                INNER JOIN Supplier ON Pembelian.id_supplier = Supplier.id_supplier
                INNER JOIN DetailBeli ON Pembelian.id_pembelian = DetailBeli.id_pembelian
                LEFT JOIN Utang ON Pembelian.id_pembelian = Utang.id_pembelian
                GROUP BY Pembelian.id_pembelian
                ''')

        rows = c.fetchall()
        for row in rows:
            id_pembelian = row[0]
            tanggal = row[1]
            nama_supplier = row[2]
            nilai_utang = row[3] if row[3] else 0
            pembayaran = row[4] if row[4] else 0
            sisa_utang = nilai_utang - pembayaran
            
            # Insert data into the table view
            treeviewutang.insert("", tk.END, values=(id_pembelian, tanggal, nama_supplier, nilai_utang, pembayaran, sisa_utang))

    # Function to generate the Piutang report
    def generate_piutang_report():
        # Clear existing table data
        for row in treeviewpiutang.get_children():
            treeviewpiutang.delete(row)

        c.execute('''SELECT Penjualan.id_penjualan, Penjualan.tanggal_jual, Penjualan.nama_pembeli,
                        DetailJual.harga_jual * DetailJual.jumlah - Penjualan.pembayaran AS total_piutang, 
                        SUM(Piutang.pembayaran)
                FROM Penjualan
                INNER JOIN DetailJual ON Penjualan.id_penjualan = DetailJual.id_penjualan
                LEFT JOIN Piutang ON Penjualan.id_penjualan = Piutang.id_penjualan
                GROUP BY Penjualan.id_penjualan
                ''')

        rows = c.fetchall()
        for row in rows:
            id_penjualan = row[0]
            tanggal = row[1]
            nama_pembeli = row[2]
            total_piutang = row[3] if row[3] else 0
            pembayaran = row[4] if row[4] else 0
            sisa_piutang = total_piutang - pembayaran

            # Insert data into the table view
            treeviewpiutang.insert("", tk.END, values=(id_penjualan, tanggal, nama_pembeli, total_piutang, pembayaran, sisa_piutang))

    # Fetch and display initial data in the table view
    generate_piutang_report()
    generate_utang_report()

def input_saldo_awal():
    # Hapus konten sebelumnya (jika ada)
    for widget in content_frame.winfo_children():
        widget.destroy()

    def save_data():
        workbook = Workbook()
        sheet = workbook.active
        
        # Get the data from the Entry widget
        saldo = entry_saldo.get()
        piutang = entry_piutang.get()
        utang = entry_utang.get()
        persediaan = entry_persediaan.get()

        # Write the data to the first cell in the worksheet
        sheet['A1'] = 'saldo'
        sheet['A2'] = saldo
        sheet['D1'] = 'piutang'
        sheet['D2'] = piutang
        sheet['E1'] = 'utang'
        sheet['E2'] = utang
        sheet['F1'] = 'persediaan'
        sheet['F2'] = persediaan
        # Save the workbook to an Excel file
        workbook.save('data.xlsx')

    label_saldo = customtkinter.CTkLabel(content_frame, text="Saldo Awal")
    entry_saldo = customtkinter.CTkEntry(content_frame)
    label_piutang = customtkinter.CTkLabel(content_frame, text="Piutang Awal")
    entry_piutang = customtkinter.CTkEntry(content_frame)
    label_utang = customtkinter.CTkLabel(content_frame, text="Utang Awal")
    entry_utang = customtkinter.CTkEntry(content_frame)
    label_persediaan = customtkinter.CTkLabel(content_frame, text="Total Persediaan Awal")
    entry_persediaan = customtkinter.CTkEntry(content_frame)

    # Position the form widgets
    label_saldo.grid(row=0, column=0, padx=20, pady=5)
    entry_saldo.grid(row=1, column=0, padx=20, pady=5)
    label_piutang.grid(row=2, column=0, padx=20, pady=5)
    entry_piutang.grid(row=3, column=0, padx=20, pady=5)
    label_utang.grid(row=4, column=0, padx=20, pady=5)
    entry_utang.grid(row=5, column=0, padx=20, pady=5)
    label_persediaan.grid(row=6, column=0, padx=20, pady=5)
    entry_persediaan.grid(row=7, column=0, padx=20, pady=5)

    button_piutang = customtkinter.CTkButton(content_frame, text="Save", command=save_data)
    button_piutang.grid(row=8, column=0, padx=5, pady=50)

def show_laporankeuangan():
    # Hapus konten sebelumnya (jika ada)
    for widget in content_frame.winfo_children():
        widget.destroy()

    # Function to calculate and display the financial report
    def generate_report():
        
        # Get the selected period from the entry fields
        start_date = entry_start_date.get()
        end_date = entry_end_date.get()

        # Calculate the total sales (penjualan) within the selected period
        c.execute('''SELECT SUM(harga_jual * jumlah) AS total FROM DetailJual
                  INNER JOIN Penjualan ON DetailJual.id_penjualan = Penjualan.id_penjualan
                  WHERE tanggal_jual BETWEEN ? AND ?''', (start_date, end_date))
        total_sales = c.fetchone()[0]
        if total_sales is None:
            total_sales = 0

        # Calculate the total sales (penjualan) within the selected period
        c.execute('''SELECT SUM(harga_beli * jumlah_produk) AS total FROM DetailBeli
                  INNER JOIN Pembelian ON DetailBeli.id_pembelian = Pembelian.id_pembelian
                  WHERE tanggal_beli BETWEEN ? AND ?''', (start_date, end_date))
        total_buy = c.fetchone()[0]
        if total_buy is None:
            total_buy = 0
        
        # Get the total operating expenses from the Pengeluaran table within the selected period
        c.execute("SELECT SUM(pembayaran) FROM Pembayaran WHERE Jenis='Pembayaran Gaji dan Upah' AND tanggal_transaksi BETWEEN ? AND ?", (start_date, end_date))
        gaji_upah = c.fetchone()[0]
        if gaji_upah is None:
            gaji_upah = 0

        # Get the total operating expenses from the Pengeluaran table within the selected period
        c.execute("SELECT SUM(pembayaran) FROM Pembayaran WHERE Jenis='Pembayaran Listrik' AND tanggal_transaksi BETWEEN ? AND ?", (start_date, end_date))
        listrik = c.fetchone()[0]
        if listrik is None:
            listrik = 0

        # Get the total operating expenses from the Pengeluaran table within the selected period
        c.execute("SELECT SUM(pembayaran) FROM Pembayaran WHERE Jenis='Pembayaran Maintenance' AND tanggal_transaksi BETWEEN ? AND ?", (start_date, end_date))
        maintenance = c.fetchone()[0]
        if maintenance is None:
            maintenance = 0

        # Get the total operating expenses from the Pengeluaran table within the selected period
        c.execute("SELECT SUM(pembayaran) FROM Pembayaran WHERE Jenis='Pembayaran Internet dan Telepon' AND tanggal_transaksi BETWEEN ? AND ?", (start_date, end_date))
        internet = c.fetchone()[0]
        if internet is None:
            internet = 0
        
        # Get the total operating expenses from the Pengeluaran table within the selected period
        c.execute("SELECT SUM(pembayaran) FROM Pembayaran WHERE Jenis='Pembayaran Bahan Bakar' AND tanggal_transaksi BETWEEN ? AND ?", (start_date, end_date))
        gas = c.fetchone()[0]
        if gas is None:
            gas = 0

        total_expenses =  gaji_upah + listrik + gas + internet + maintenance

        workbook = load_workbook('data.xlsx')
        sheet = workbook.active
        persediaan_awal = float(sheet['F2'].value)

        c.execute("SELECT id_produk FROM Produk")
        produk_rows = c.fetchall()

        inventory = []

        for row in produk_rows:
            produk_id = row[0]

            # Calculate total quantity produced
            c.execute("SELECT SUM(harga_beli * jumlah_produk) FROM DetailBeli WHERE id_produk=?", (produk_id,))
            total_buy = c.fetchone()[0]

            # Calculate total quantity sold
            c.execute("SELECT SUM(harga_jual * jumlah) FROM DetailJual WHERE id_produk=?", (produk_id,))
            total_sold = c.fetchone()[0]

            # Calculate the remaining quantity in inventory
            if total_buy is None:
                total_buy = 0
            if total_sold is None:
                total_sold = 0

            remaining_value = total_buy - total_sold
            inventory.append(remaining_value)

        persediaan_akhir = sum(inventory)

        harga_pokok_penjualan = total_buy + persediaan_awal - persediaan_akhir
        
        # Calculate the gross profit
        gross_profit = total_sales - harga_pokok_penjualan

        # Calculate the net profit
        net_profit = gross_profit - total_expenses

        workbook = load_workbook('data.xlsx')
        sheet = workbook.active
        
        # Write the data to the first cell in the worksheet
        sheet['C1'] = 'laba bersih'
        sheet['C2'] = net_profit
        
        # Save the workbook to an Excel file
        workbook.save('data.xlsx')
        
        # Insert data into the table view for the income statement
        # Insert data into the table view for the income statement
        treeview.insert("", tk.END, values=("Pendapatan"))
        treeview.insert("", tk.END, values=("Penjualan", format_currency(total_sales)))
        treeview.insert("", tk.END, values=("Harga Pokok Penjualan", format_currency(harga_pokok_penjualan)))
        treeview.insert("", tk.END, values=("Laba Kotor", "", format_currency(gross_profit)))
        treeview.insert("", tk.END, values=(" "))
        treeview.insert("", tk.END, values=("Beban Operasional",""))
        treeview.insert("", tk.END, values=("Pembayaran Gaji dan Upah", format_currency(gaji_upah)))
        treeview.insert("", tk.END, values=("Pembayaran Listrik", format_currency(listrik)))
        treeview.insert("", tk.END, values=("Pembayaran Bahan Bakar", format_currency(gas)))
        treeview.insert("", tk.END, values=("Pembayaran Internet dan Telepon", format_currency(internet)))
        treeview.insert("", tk.END, values=("Pembayaran Maintenance", format_currency(maintenance)))
        treeview.insert("", tk.END, values=("Total Beban Operasional", "", format_currency(total_expenses)))
        treeview.insert("", tk.END, values=(""))
        treeview.insert("", tk.END, values=("Laba/(Rugi) Sebelum Pajak", format_currency(gross_profit)))
        treeview.insert("", tk.END, values=("Laba Bersih", format_currency(net_profit)))

    label_start_date = customtkinter.CTkLabel(content_frame, text="Start Date")
    entry_start_date = DateEntry(content_frame, width=23, background='darkblue', foreground='white', date_pattern='yyyy-mm-dd')

    label_end_date = customtkinter.CTkLabel(content_frame, text="End Date")
    entry_end_date = DateEntry(content_frame, width=23, background='darkblue', foreground='white', date_pattern='yyyy-mm-dd')

    button_generate = customtkinter.CTkButton(content_frame, text="Generate Report", command=generate_report)

    label_start_date.grid(row=0, column=0, padx=5, pady=5)
    entry_start_date.grid(row=0, column=1, padx=5, pady=5)

    label_end_date.grid(row=1, column=0, padx=5, pady=5)
    entry_end_date.grid(row=1, column=1, padx=5, pady=5)

    button_generate.grid(row=2, columnspan=2, padx=5, pady=5)

    treeview = ttk.Treeview(content_frame, columns=("Deskripsi", "", ""), show="headings", height=20)
    treeview.heading("Deskripsi", text="Deskripsi")
    treeview.heading("", text="")
    treeview.heading("", text="")

    treeview.column("Deskripsi", width=200)
    treeview.column("", width=150)
    treeview.column("", width=150)

    treeview.grid(row=3, column=1, columnspan=2, padx=5, pady=5,)

def show_laporanaruskas():

    # Hapus konten sebelumnya (jika ada)
    for widget in content_frame.winfo_children():
        widget.destroy()

    # Function to calculate and display the financial report
    def generate_report():
        
        # Get the selected period from the entry fields
        start_date = entry_start_date.get()
        end_date = entry_end_date.get()
        
        # Calculate the total sales (penjualan) within the selected period
        c.execute("SELECT SUM(pembayaran) FROM Penjualan WHERE tanggal_jual BETWEEN ? AND ?", (start_date, end_date))
        total_sales = c.fetchone()[0]
        if total_sales is None:
            total_sales = 0

        c.execute("SELECT SUM(pembayaran) FROM Piutang WHERE tanggal_bayar BETWEEN ? AND ?", (start_date, end_date))
        penerimaan_piutang = c.fetchone()[0]
        if penerimaan_piutang is None:
            penerimaan_piutang = 0

        # Calculate the total sales (penjualan) within the selected period
        c.execute("SELECT SUM(pembayaran) FROM Transaksi WHERE jenis='Penerimaan Pinjaman' AND tanggal_transaksi BETWEEN ? AND ?", (start_date, end_date))
        penerimaan_pinjaman = c.fetchone()[0]
        if penerimaan_pinjaman is None:
            penerimaan_pinjaman = 0

        # Calculate the total sales (penjualan) within the selected period
        c.execute("SELECT SUM(pembayaran) FROM Transaksi WHERE jenis='Penerimaan Kas Lainnya' AND tanggal_transaksi BETWEEN ? AND ?", (start_date, end_date))
        penerimaan_kas_lainnya = c.fetchone()[0]
        if penerimaan_kas_lainnya is None:
            penerimaan_kas_lainnya = 0
        
        # Calculate the total cost of goods sold (biaya bahan baku) within the selected period
        c.execute("SELECT SUM(pembayaran) FROM Pembelian WHERE Pembelian.tanggal_beli BETWEEN ? AND ?", (start_date, end_date))
        total_cost_of_goods = c.fetchone()[0]
        if total_cost_of_goods is None:
            total_cost_of_goods = 0
        
        # Get the total operating expenses from the Pengeluaran table within the selected period
        c.execute("SELECT SUM(pembayaran) FROM Pembayaran WHERE Jenis='Pembayaran Gaji dan Upah' AND tanggal_transaksi BETWEEN ? AND ?", (start_date, end_date))
        gaji_upah = c.fetchone()[0]
        if gaji_upah is None:
            gaji_upah = 0

        # Get the total operating expenses from the Pengeluaran table within the selected period
        c.execute("SELECT SUM(pembayaran) FROM Pembayaran WHERE Jenis='Pembayaran Listrik' AND tanggal_transaksi BETWEEN ? AND ?", (start_date, end_date))
        listrik = c.fetchone()[0]
        if listrik is None:
            listrik = 0

        # Get the total operating expenses from the Pengeluaran table within the selected period
        c.execute("SELECT SUM(pembayaran) FROM Pembayaran WHERE Jenis='Pembayaran Maintenance' AND tanggal_transaksi BETWEEN ? AND ?", (start_date, end_date))
        maintenance = c.fetchone()[0]
        if maintenance is None:
            maintenance = 0

        # Get the total operating expenses from the Pengeluaran table within the selected period
        c.execute("SELECT SUM(pembayaran) FROM Pembayaran WHERE Jenis='Pembayaran Internet dan Telepon' AND tanggal_transaksi BETWEEN ? AND ?", (start_date, end_date))
        internet = c.fetchone()[0]
        if internet is None:
            internet = 0
        
        # Get the total operating expenses from the Pengeluaran table within the selected period
        c.execute("SELECT SUM(pembayaran) FROM Pembayaran WHERE Jenis='Pembayaran Bahan Bakar' AND tanggal_transaksi BETWEEN ? AND ?", (start_date, end_date))
        gas = c.fetchone()[0]
        if gas is None:
            gas = 0

        total_expenses =  gaji_upah + listrik + gas + internet + maintenance

        c.execute("SELECT SUM(pembayaran) FROM Utang WHERE tanggal_bayar BETWEEN ? AND ?", (start_date, end_date))
        bayar_utang = c.fetchone()[0]
        if bayar_utang is None:
            bayar_utang = 0

        # Get the total operating expenses from the Pengeluaran table within the selected period
        c.execute("SELECT SUM(pembayaran) FROM Transaksi WHERE Jenis='Pemberian Pinjaman' AND tanggal_transaksi BETWEEN ? AND ?", (start_date, end_date))
        pemberian_pinjaman = c.fetchone()[0]
        if pemberian_pinjaman is None:
            pemberian_pinjaman = 0

        # Get the total operating expenses from the Pengeluaran table within the selected period
        c.execute("SELECT SUM(pembayaran) FROM Transaksi WHERE Jenis='Pengeluaran Kas Lainnya' AND tanggal_transaksi BETWEEN ? AND ?", (start_date, end_date))
        pengeluaran_kas_lainnya = c.fetchone()[0]
        if pengeluaran_kas_lainnya is None:
            pengeluaran_kas_lainnya = 0

        # Get the total operating expenses from the Pengeluaran table within the selected period
        c.execute("SELECT SUM(pembayaran) FROM Transaksi WHERE Jenis='Pengambilan Pribadi' AND tanggal_transaksi BETWEEN ? AND ?", (start_date, end_date))
        pengambilan_pribadi = c.fetchone()[0]
        if pengambilan_pribadi is None:
            pengambilan_pribadi = 0

        c.execute("SELECT SUM(pembayaran) FROM Transaksi WHERE jenis='Tambahan Modal' AND tanggal_transaksi BETWEEN ? AND ?", (start_date, end_date))
        tambahan_modal = c.fetchone()[0]
        if tambahan_modal is None:
            tambahan_modal = 0

        total_outflow = total_cost_of_goods + total_expenses + bayar_utang + pemberian_pinjaman + pengeluaran_kas_lainnya

        total_inflow = total_sales + penerimaan_pinjaman + penerimaan_kas_lainnya + penerimaan_piutang + tambahan_modal

        workbook = load_workbook('data.xlsx')
        sheet = workbook.active
        saldo_awal = float(sheet['A2'].value)    
                           
        pertambahan_pengurangan_kas = total_inflow - total_outflow

        saldo_kas_akhir = saldo_awal + pertambahan_pengurangan_kas

        workbook = load_workbook('data.xlsx')
        sheet = workbook.active
        sheet['B1'] = 'saldo akhir'
        sheet['B2'] = saldo_kas_akhir
        workbook.save('data.xlsx')

        treeview.insert("", tk.END, values=("Kas Masuk",""))
        # Insert data into the table view for the income statement
        treeview.insert("", tk.END, values=("Penjualan", format_currency(total_sales)))
        treeview.insert("", tk.END, values=("Penerimaan Piutang", format_currency(penerimaan_piutang)))
        treeview.insert("", tk.END, values=("Penerimaan Pinjaman", format_currency(penerimaan_pinjaman)))
        treeview.insert("", tk.END, values=("Tambahan Modal", format_currency(tambahan_modal)))
        treeview.insert("", tk.END, values=("Penerimaan Kas Lainnya", format_currency(penerimaan_kas_lainnya)))
        treeview.insert("", tk.END, values=("Total Kas Masuk","", format_currency(total_inflow)))
        # Insert data into the table view for the income statement
        treeview.insert("", tk.END, values=(""))
        treeview.insert("", tk.END, values=("Kas Keluar",""))
        # Insert data into the table view for the balance sheet

        treeview.insert("", tk.END, values=("Pembelian Bahan Baku",format_currency(total_cost_of_goods)))
        treeview.insert("", tk.END, values=("Beban Operasional", format_currency(total_expenses)))
        treeview.insert("", tk.END, values=("Pembayaran Utang", format_currency(bayar_utang)))
        treeview.insert("", tk.END, values=("Pemberian Pinjaman", format_currency(pemberian_pinjaman)))
        treeview.insert("", tk.END, values=("Pengeluaran Kas Lainnya", format_currency(pengeluaran_kas_lainnya)))
        treeview.insert("", tk.END, values=("Total Kas Keluar", "", format_currency(total_outflow)))
        treeview.insert("", tk.END, values=(" "))
        treeview.insert("", tk.END, values=("Saldo Kas Awal",format_currency(saldo_awal)))
        treeview.insert("", tk.END, values=("Pertambahan/Pengurangan Kas","", format_currency(total_inflow-total_outflow)))
        treeview.insert("", tk.END, values=("Saldo Kas Akhir","", format_currency(saldo_kas_akhir)))
    
    label_start_date = customtkinter.CTkLabel(content_frame, text="Start Date")
    entry_start_date = DateEntry(content_frame, width=23, background='darkblue', foreground='white', date_pattern='yyyy-mm-dd')

    label_end_date = customtkinter.CTkLabel(content_frame, text="End Date")
    entry_end_date = DateEntry(content_frame, width=23, background='darkblue', foreground='white', date_pattern='yyyy-mm-dd')

    button_generate = customtkinter.CTkButton(content_frame, text="Generate Report", command=generate_report)

    label_start_date.grid(row=0, column=0, padx=5, pady=5)
    entry_start_date.grid(row=0, column=1, padx=5, pady=5)

    label_end_date.grid(row=1, column=0, padx=5, pady=5)
    entry_end_date.grid(row=1, column=1, padx=5, pady=5)

    button_generate.grid(row=2, column=1, padx=5, pady=5)

    treeview = ttk.Treeview(content_frame, columns=("Deskripsi", "", ""), show="headings", height=20)
    treeview.heading("Deskripsi", text="Deskripsi")
    treeview.heading("", text="")
    treeview.heading("", text="")

    treeview.column("Deskripsi", width=200)
    treeview.column("", width=50)
    treeview.column("", width=200)

    treeview.grid(row=3, column=1, columnspan=3, padx=5, pady=5)

def show_laporanneraca():
        # Hapus konten sebelumnya (jika ada)
    for widget in content_frame.winfo_children():
        widget.destroy()

    # Function to calculate and display the financial report
    def generate_report():
        
        # Get the selected period from the entry fields
        start_date = entry_start_date.get()
        end_date = entry_end_date.get()

        # Load the Excel file
        workbook = load_workbook('data.xlsx')
        sheet = workbook.active

        modal_awal = sheet['A2'].value 
        total_kas = sheet['B2'].value
        laba_bersih = sheet['C2'].value
        piutang_awal = sheet['D2'].value
        utang_awal = sheet['E2'].value

        c.execute("SELECT id_produk FROM Produk")
        produk_rows = c.fetchall()

        inventory = []

        for row in produk_rows:
            produk_id = row[0]

            # Calculate total quantity produced
            c.execute("SELECT SUM(harga_beli * jumlah_produk) FROM DetailBeli WHERE id_produk=?", (produk_id,))
            total_buy = c.fetchone()[0]

            # Calculate total quantity sold
            c.execute("SELECT SUM(harga_jual * jumlah) FROM DetailJual WHERE id_produk=?", (produk_id,))
            total_sold = c.fetchone()[0]

            # Calculate the remaining quantity in inventory
            if total_buy is None:
                total_buy = 0
            if total_sold is None:
                total_sold = 0

            remaining_value = total_buy - total_sold
            inventory.append(remaining_value)

        total_inventory_value = sum(inventory)
        
        c.execute('''SELECT DetailJual.harga_jual * DetailJual.jumlah - Penjualan.pembayaran AS total_piutang, 
                        SUM(Piutang.pembayaran)
                FROM Penjualan
                INNER JOIN DetailJual ON Penjualan.id_penjualan = DetailJual.id_penjualan
                LEFT JOIN Piutang ON Penjualan.id_penjualan = Piutang.id_penjualan
                GROUP BY Penjualan.id_penjualan
                ''')

        rows = c.fetchall()
        piutang = []
        for row in rows:
            total_piutang = row[0] if row[0] else 0
            pembayaran = row[1] if row[1] else 0
            sisa_piutang = total_piutang - pembayaran
            piutang.append(sisa_piutang)

        total_accounts_receivable = float(piutang_awal) + sum(piutang)

        c.execute('''SELECT DetailBeli.harga_beli * DetailBeli.jumlah_produk - Pembelian.pembayaran AS total_utang, 
                        SUM(Utang.pembayaran)
                FROM Pembelian
                INNER JOIN DetailBeli ON Pembelian.id_pembelian = DetailBeli.id_pembelian
                LEFT JOIN Utang ON Pembelian.id_pembelian = Utang.id_pembelian
                GROUP BY Pembelian.id_pembelian
                ''')

        rows = c.fetchall()
        utang = []
        for row in rows:
            nilai_utang = row[0] if row[0] else 0
            pembayaran = row[1] if row[1] else 0
            sisa_utang = nilai_utang - pembayaran
            utang.append(sisa_utang)

        total_accounts_payable = float(utang_awal) + sum(utang)

        total_asset = total_kas + total_inventory_value + total_accounts_receivable

        # Calculate the total sales (penjualan) within the selected period
        c.execute("SELECT SUM(pembayaran) FROM Transaksi WHERE jenis='Penerimaan Kas Lainnya' AND tanggal_transaksi BETWEEN ? AND ?", (start_date, end_date))
        penerimaan_kas_lainnya = c.fetchone()[0]
        if penerimaan_kas_lainnya is None:
            penerimaan_kas_lainnya = 0

        # Get the total operating expenses from the Pengeluaran table within the selected period
        c.execute("SELECT SUM(pembayaran) FROM Transaksi WHERE Jenis='Pengeluaran Kas Lainnya' AND tanggal_transaksi BETWEEN ? AND ?", (start_date, end_date))
        pengeluaran_kas_lainnya = c.fetchone()[0]
        if pengeluaran_kas_lainnya is None:
            pengeluaran_kas_lainnya = 0

        # Get the total operating expenses from the Pengeluaran table within the selected period
        c.execute("SELECT SUM(pembayaran) FROM Transaksi WHERE Jenis='Pengambilan Pribadi' AND tanggal_transaksi BETWEEN ? AND ?", (start_date, end_date))
        pengambilan_pribadi = c.fetchone()[0]
        if pengambilan_pribadi is None:
            pengambilan_pribadi = 0

        c.execute("SELECT SUM(pembayaran) FROM Transaksi WHERE jenis='Tambahan Modal' AND tanggal_transaksi BETWEEN ? AND ?", (start_date, end_date))
        tambahan_modal = c.fetchone()[0]
        if tambahan_modal is None:
            tambahan_modal = 0

        laba_ditahan = float(laba_bersih) + penerimaan_kas_lainnya - pengambilan_pribadi - pengeluaran_kas_lainnya

        total_modal = float(modal_awal) + laba_ditahan + tambahan_modal

        total_kewajiban_modal = total_modal + total_accounts_payable

        treeview.insert("", tk.END, values=("Aset", ""))
        treeview.insert("", tk.END, values=("Total Kas",format_currency(float(total_kas))))
        treeview.insert("", tk.END, values=("Persediaan Produk", format_currency(total_inventory_value)))
        treeview.insert("", tk.END, values=("Piutang", format_currency(total_accounts_receivable)))
        treeview.insert("", tk.END, values=("Total Aset","", format_currency(total_asset)))
        treeview.insert("", tk.END, values=("", ""))
        treeview.insert("", tk.END, values=("Kewajiban + Modal", ""))
        treeview.insert("", tk.END, values=("Utang", format_currency(total_accounts_payable)))
        treeview.insert("", tk.END, values=("Total Kewajiban","", format_currency(total_accounts_payable)))
        treeview.insert("", tk.END, values=("", ""))
        treeview.insert("", tk.END, values=("Modal",""))
        treeview.insert("", tk.END, values=("Modal Awal", format_currency(float(modal_awal))))
        treeview.insert("", tk.END, values=("Laba Ditahan", format_currency(laba_ditahan)))
        treeview.insert("", tk.END, values=("Tambahan Modal", format_currency(tambahan_modal)))
        treeview.insert("", tk.END, values=("Total Modal","", format_currency(total_modal)))
        treeview.insert("", tk.END, values=("", ""))
        treeview.insert("", tk.END, values=("Total Kewajiban + Modal", "", format_currency(total_kewajiban_modal)))

    label_start_date = customtkinter.CTkLabel(content_frame, text="Start Date")
    entry_start_date = DateEntry(content_frame, width=23, background='darkblue', foreground='white', date_pattern='yyyy-mm-dd')

    label_end_date = customtkinter.CTkLabel(content_frame, text="End Date")
    entry_end_date = DateEntry(content_frame, width=23, background='darkblue', foreground='white', date_pattern='yyyy-mm-dd')

    button_generate = customtkinter.CTkButton(content_frame, text="Generate Report", command=generate_report)

    label_start_date.grid(row=0, column=0, padx=5, pady=5)
    entry_start_date.grid(row=0, column=1, padx=5, pady=5)

    label_end_date.grid(row=1, column=0, padx=5, pady=5)
    entry_end_date.grid(row=1, column=1, padx=5, pady=5)

    button_generate.grid(row=2, columnspan=2, padx=5, pady=5)

    treeview = ttk.Treeview(content_frame, columns=("Deskripsi", "",""), show="headings", height=20)
    treeview.heading("Deskripsi", text="Deskripsi")
    treeview.heading("", text="")
    treeview.heading("", text="")

    treeview.column("Deskripsi", width=200)
    treeview.column("", width=50)
    treeview.column("", width=230)

    treeview.grid(row=3, column=1, columnspan=3, padx=5, pady=5,)

def show_pengaturan():
    # Hapus konten sebelumnya (jika ada)
    for widget in content_frame.winfo_children():
        widget.destroy()

    # Tambahkan tombol Produk
    produk_button = customtkinter.CTkButton(content_frame, text="Produk", command=show_produk, font=("Arial",30), height=70, width=400)
    produk_button.pack(pady=80, padx=160)

    # Tambahkan tombol Supplier
    supplier_button = customtkinter.CTkButton(content_frame, text="Supplier", command=show_supplier, font=("Arial",30), height=70, width=400)
    supplier_button.pack(pady=45, padx=160)

    # Tambahkan tombol Supplier
    supplier_button = customtkinter.CTkButton(content_frame, text="Pengaturan Awal", command=input_saldo_awal, font=("Arial",30), height=70, width=400)
    supplier_button.pack(pady=75, padx=160)

def show_laporan():
    # Hapus konten sebelumnya (jika ada)
    for widget in content_frame.winfo_children():
        widget.destroy()

    # Tambahkan tombol
    laporanstok_button = customtkinter.CTkButton(content_frame, text="Informasi Stok Produk", command=show_persediaanproduk, font=("Arial",30), height=60, width=500)
    laporanstok_button.pack(pady=19, padx=130)

    # Tambahkan tombol
    laporanpenjualan_button = customtkinter.CTkButton(content_frame, text="Laporan Penjualan", command=show_laporanpenjualan, font=("Arial",30), height=60, width=400)
    laporanpenjualan_button.pack(pady=19, padx=130)

    # Tambahkan tombol
    laporanpembelian_button = customtkinter.CTkButton(content_frame, text="Laporan Pembelian", command=show_laporanpembelian, font=("Arial",30), height=60, width=400)
    laporanpembelian_button.pack(pady=18, padx=130)

    # Tambahkan tombol
    laporanutangpiutang_button = customtkinter.CTkButton(content_frame, text="Laporan Utang Piutang", command=show_laporanutangpiutang, font=("Arial",30), height=60, width=400)
    laporanutangpiutang_button.pack(pady=19, padx=130)

    # Tambahkan tombol
    laporanaruskas_button = customtkinter.CTkButton(content_frame, text="Laporan Arus Kas", command=show_laporanaruskas, font=("Arial",30), height=60, width=400)
    laporanaruskas_button.pack(pady=18, padx=130)

    # Tambahkan tombol
    laporanlabarugi_button = customtkinter.CTkButton(content_frame, text="Laporan Laba Rugi", command=show_laporankeuangan, font=("Arial",30), height=60, width=400)
    laporanlabarugi_button.pack(pady=19, padx=130)

    # Tambahkan tombol
    laporanneraca_button = customtkinter.CTkButton(content_frame, text="Laporan Neraca", command=show_laporanneraca, font=("Arial",30), height=60, width=400)
    laporanneraca_button.pack(pady=18, padx=130)

root = customtkinter.CTk(fg_color="#FF8C52")
root.iconbitmap('C:/Users/Reza/Downloads/sambal4')
root.title("UD Sriwijaya Karangploso")

# Mendapatkan lebar dan tinggi jendela
root_width = 920
root_height = 700

# Mengatur jendela agar berada di tengah layar
center_window(root, root_width, root_height)

# Buat frame untuk menu navigasi
menu_frame = tk.Frame(root, background="#FEF4DF")
menu_frame.pack(side="left", fill="y")

# load logo
file_path = os.path.dirname(os.path.realpath(__file__))
image_dashboard = customtkinter.CTkImage(Image.open(file_path + "/Logotype.png"), size=(120,55))

# Tombol Dashboard
dashboard_button = customtkinter.CTkButton(menu_frame, text="", image=image_dashboard, command=show_dashboard,  fg_color="#FEF4DF", hover_color="#e5dcc9", height=70)
dashboard_button.pack(padx=5, pady=5)

# Tombol Pembelian
pembelian_button = customtkinter.CTkButton(menu_frame, text="Pembelian", command=show_pembelian, fg_color="#FF8C52", hover_color="#72C822", height=50, font=("Arial",20))
pembelian_button.pack(padx=5, pady=5)

# Tombol Penjualan
penjualan_button = customtkinter.CTkButton(menu_frame, text="Penjualan", command=show_penjualan, fg_color="#FF8C52", hover_color="#72C822", height=50, font=("Arial",20))
penjualan_button.pack(padx=5, pady=5)

# Tombol Pengeluaran
pembayaran_button = customtkinter.CTkButton(menu_frame, text="Pembayaran", command=show_pembayaran, fg_color="#FF8C52", hover_color="#72C822", height=50, font=("Arial",20))
pembayaran_button.pack(padx=5, pady=5)

# Tombol Penerimaan
transaksi_button = customtkinter.CTkButton(menu_frame, text="Transaksi", command=show_transaksilain, fg_color="#FF8C52", hover_color="#72C822", height=50, font=("Arial",20))
transaksi_button.pack(padx=5, pady=5)

# Tombol Piutang
piutang_button = customtkinter.CTkButton(menu_frame, text="Piutang", command=show_piutang, fg_color="#FF8C52", hover_color="#72C822", height=50, font=("Arial",20))
piutang_button.pack(padx=5, pady=5)

# Tombol Utang
utang_button = customtkinter.CTkButton(menu_frame, text="Utang", command=show_utang, fg_color="#FF8C52", hover_color="#72C822", height=50, font=("Arial",20))
utang_button.pack(padx=5, pady=5)

# Tombol Pengaturan
pengaturan_button = customtkinter.CTkButton(menu_frame, text="Pengaturan", command=show_pengaturan, fg_color="#FF8C52", hover_color="#72C822", height=50, font=("Arial",20))
pengaturan_button.pack(padx=5, pady=5)

# Tombol Laporan
pengaturan_button = customtkinter.CTkButton(menu_frame, text="Laporan", command=show_laporan, fg_color="#FF8C52", hover_color="#72C822", height=50, font=("Arial",20))
pengaturan_button.pack(padx=5, pady=5)

# Create a canvas widget
canvas = tk.Canvas(root, background="#FEF4DF")
canvas.pack(side="left", fill="both", expand=True)

# Create a scrollbar widget
scrollbar = ttk.Scrollbar(root, orient="vertical", command=canvas.yview)
scrollbar.pack(side="right", fill="y")

# Configure the canvas to use the scrollbar
canvas.configure(yscrollcommand=scrollbar.set)
canvas.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

# Buat frame untuk konten
content_frame = customtkinter.CTkFrame(master=canvas, corner_radius=20, fg_color="#FEF4DF")
canvas.create_window((0, 0), window=content_frame, anchor="nw")

def update_scroll_region(event):
    canvas.configure(scrollregion=canvas.bbox("all"))

content_frame.bind("<Configure>", update_scroll_region)

# Pack the canvas and scrollbar to the root window
canvas.pack(side="left", fill="both", expand=True)
scrollbar.pack(side="right", fill="y")

# Tampilkan konten dashboard awal
show_dashboard()

# Jalankan aplikasi
root.mainloop()